"""
employees_module.py — إدارة ملفات الموظفين + PDF عربي عبر wkhtmltopdf
"""
import streamlit as st
import pandas as pd
import json, io, os, subprocess, tempfile, base64
from datetime import date

EMP_FILE = "emp_profiles.json"
_BASE = os.path.dirname(os.path.abspath(__file__))

# ─────────────────────────────
# Fonts
# ─────────────────────────────
def _load_font_b64(fn):
    p = os.path.join(_BASE, fn)
    if os.path.exists(p):
        with open(p,"rb") as f:
            return base64.b64encode(f.read()).decode()
    return ""

_FONT_B64      = _load_font_b64("Amiri-Regular.ttf") or _load_font_b64("DejaVuSans.ttf")
_FONT_BOLD_B64 = _load_font_b64("Amiri-Bold.ttf")    or _load_font_b64("DejaVuSans-Bold.ttf")


# ─────────────────────────────
# WKHTMLTOPDF
# ─────────────────────────────
def _find_wk():
    for c in ["wkhtmltopdf",
              r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe",
              r"C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe",
              "/usr/bin/wkhtmltopdf","/usr/local/bin/wkhtmltopdf"]:
        try:
            if subprocess.run([c,"--version"],capture_output=True,timeout=5).returncode==0:
                return c
        except:
            pass
    return None

_WK = _find_wk()


# ─────────────────────────────
# DATA IO
# ─────────────────────────────
def load_profiles():
    if os.path.exists(EMP_FILE):
        with open(EMP_FILE,"r",encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_profiles(p):
    with open(EMP_FILE,"w",encoding="utf-8") as f:
        json.dump(p,f,ensure_ascii=False,indent=2)


# ─────────────────────────────
# Helpers
# ─────────────────────────────
def verbal(pct):
    try:
        p=float(pct)
    except:
        return "---"
    return "ممتاز" if p>=90 else "جيد جدا" if p>=80 else "جيد" if p>=70 else "متوسط" if p>=60 else "ضعيف"

def vclr(pct):
    try:
        p=float(pct)
    except:
        return "#64748B"
    return "#15803d" if p>=80 else "#b45309" if p>=60 else "#b91c1c"


def safe_sheet_name(name):
    return "".join(c for c in str(name) if c not in r'[]:*?/\\').strip()[:28]


# ─────────────────────────────
# CSS PDF
# ─────────────────────────────
_CSS = f"""
@font-face{{font-family:'AF';font-weight:400;src:url('data:font/truetype;base64,{_FONT_B64}') format('truetype');}}
@font-face{{font-family:'AF';font-weight:700;src:url('data:font/truetype;base64,{_FONT_BOLD_B64}') format('truetype');}}
*{{font-family:'AF',Arial,sans-serif;direction:rtl;box-sizing:border-box;margin:0;padding:0;}}
body{{font-size:12px;color:#1a1a1a;padding:10mm 15mm;line-height:1.6;}}
"""


# ─────────────────────────────
# BODY HTML (PDF)
# ─────────────────────────────
def _body(profile, df_data):
    nm=profile.get("name","")
    eid=profile.get("emp_id","---")

    b=f"<h2>{nm}</h2>"
    b+=f"<p>{eid}</p>"
    return b


def _html(profile, df_data, app_cfg, logo_path):
    co=app_cfg.get("company_name","مجموعة شركات فنون")
    body=_body(profile,df_data)

    return f"""
<!DOCTYPE html>
<html dir="rtl">
<head><meta charset="utf-8"><style>{_CSS}</style></head>
<body>
<h1>{co}</h1>
{body}
</body>
</html>
"""


# ─────────────────────────────
# PDF GENERATOR
# ─────────────────────────────
def _to_pdf(html):
    if not _WK:
        return None,"wkhtmltopdf غير مثبت"

    with tempfile.TemporaryDirectory() as tmp:
        hf=os.path.join(tmp,"cv.html")
        pf=os.path.join(tmp,"cv.pdf")

        with open(hf,"w",encoding="utf-8") as f:
            f.write(html)

        r=subprocess.run([_WK,"--encoding","utf-8","--page-size","A4",
            "--margin-top","8mm","--margin-bottom","8mm",
            "--margin-left","5mm","--margin-right","5mm",
            "--enable-local-file-access","--quiet",hf,pf],
            capture_output=True,timeout=30)

        if os.path.exists(pf) and os.path.getsize(pf)>1000:
            with open(pf,"rb") as f:
                return f.read(),None

        return None,r.stderr.decode(errors="ignore")[:200]


# ─────────────────────────────
# BUILD PDF
# ─────────────────────────────
def build_cv_pdf(eid,profile,df_data,df_kpi,app_cfg,logo_path):
    pdf,err=_to_pdf(_html(profile,df_data,app_cfg,logo_path))
    return io.BytesIO(pdf) if pdf else (None,err)


# ─────────────────────────────
# EMPLOYEES LIST RENDER
# ─────────────────────────────
def render_employee_management(df_emp,df_data,df_kpi,app_cfg,logo_path):
    profiles=load_profiles()

    st.markdown("#### 📋 إدارة الموظفين")

    all_n=sorted(set(
        [str(n).strip() for n in df_emp["EmployeeName"].dropna()
         if str(n).strip() not in ("","nan")] + list(profiles.keys())
    ))

    if not all_n:
        st.info("لا يوجد موظفون.")
        return

    for name in all_n:
        st.write(name)


# ─────────────────────────────
# EXCEL REPORT (FIXED ONLY)
# ─────────────────────────────
def _build_cv_sheet(wb, name, df_emp, df_data, df_kpi, profiles, inc_monthly, year, app_cfg=None):
    """يبني شيت CV كامل للموظف."""
    if app_cfg is None: app_cfg = {}
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(safe_sheet_name(name))
    ws.sheet_view.rightToLeft  = True
    ws.sheet_view.showGridLines = False

    # ── ألوان ───────────────────────────────────────────────────────
    DARK   = "1F3864"; MID  = "2E75B6"; ORANGE = "ED7D31"
    LGRAY  = "F2F2F2"; WHITE= "FFFFFF"; GREEN  = "E2EFDA"
    YELLOW = "FFF2CC"; RED  = "FFDAD9"; LBLUE  = "BDD7EE"
    CREAM  = "FFFBF0"

    def fill(c): return PatternFill("solid", fgColor=c.lstrip("#"))
    def fnt(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color.lstrip("#"), size=size, name="Arial")
    def aln(h="right", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrapText=wrap, readingOrder=2)

    thin = Side(style="thin",   color="AAAAAA")
    med  = Side(style="medium", color="1F3864")
    IB = Border(left=thin, right=thin, top=thin,  bottom=thin)
    OB = Border(left=med,  right=med,  top=med,   bottom=med)

    def sc(cell, val=None, bold=False, sz=10, color="000000",
           bg=None, ah="right", av="center", wrap=False, brd="inner"):
        if val is not None: cell.value = val
        cell.font      = Font(bold=bold, size=sz, color=color.lstrip("#"), name="Arial")
        cell.alignment = Alignment(horizontal=ah, vertical=av,
                                   wrapText=wrap, readingOrder=2)
        if bg: cell.fill = PatternFill("solid", fgColor=bg.lstrip("#"))
        if   brd == "outer": cell.border = OB
        elif brd == "inner": cell.border = IB

    def mc(r1,c1,r2,c2, val=None, **kw):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        sc(ws.cell(r1,c1,val), **kw)

    # ── عرض الأعمدة ─────────────────────────────────────────────────
    for col,w in [("A",28),("B",18),("C",14),("D",14),("E",14),("F",14)]:
        ws.column_dimensions[col].width = w

    # ── بيانات الموظف من df_emp ──────────────────────────────────────
    ei = df_emp[df_emp["EmployeeName"] == name]
    job_title = str(ei.iloc[0,1]).strip() if not ei.empty else "—"
    dept      = str(ei.iloc[0,2]).strip() if not ei.empty else "—"
    mgr       = str(ei.iloc[0,3]).strip() if not ei.empty else "—"
    prof      = profiles.get(name, {})

    r = 1

    # ── ترويسة ──────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 30
    mc(r,1,r,6, f"ملف الموظف — {name}",
       bold=True, sz=14, color=WHITE, bg=DARK, ah="center", brd="outer")
    r += 1

    ws.row_dimensions[r].height = 18
    mc(r,1,r,6, f"نظام تقييم الأداء — {app_cfg.get('company_name','مجموعة شركات فنون')} | {date.today().strftime('%d/%m/%Y')}",
       bold=False, sz=9, color=WHITE, bg=MID, ah="center", brd="inner")
    r += 1; r += 1

    # ── بيانات أساسية ────────────────────────────────────────────────
    ws.row_dimensions[r].height = 18
    mc(r,1,r,6, "📋 البيانات الأساسية",
       bold=True, sz=11, color=WHITE, bg=MID, ah="right", brd="outer")
    r += 1

    info_rows = [
        ("الاسم الكامل",      name),
        ("المسمى الوظيفي",    job_title),
        ("القسم",             dept),
        ("المدير المباشر",    mgr),
        ("تاريخ التعيين",     prof.get("hire_date", "—")),
        ("رقم الهاتف",        prof.get("phone", "—")),
        ("البريد الإلكتروني", prof.get("email", "—")),
        ("الجنسية",           prof.get("nationality", "—")),
    ]
    for label, val in info_rows:
        ws.row_dimensions[r].height = 16
        sc(ws.cell(r,1,label), bold=True, bg=LBLUE, ah="right", brd="inner")
        mc(r,2,r,6, val, bold=False, bg=CREAM, ah="right", brd="inner")
        r += 1
    r += 1

    # ── ملخص الأداء السنوي ──────────────────────────────────────────
    MONTHS_EN    = ["January","February","March","April","May","June",
                    "July","August","September","October","November","December"]
    MONTHS_AR    = ["يناير","فبراير","مارس","أبريل","مايو","يونيو",
                    "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
    MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]

    if df_data is not None and not df_data.empty and "EmployeeName" in df_data.columns:
        emp_data = df_data[
            (df_data["EmployeeName"] == name) &
            (df_data["Year"].astype(str) == str(year))
        ]
    else:
        emp_data = None

    ws.row_dimensions[r].height = 18
    mc(r,1,r,6, f"📊 ملخص الأداء السنوي — {year}",
       bold=True, sz=11, color=WHITE, bg=ORANGE, ah="right", brd="outer")
    r += 1

    # رأس جدول الأشهر
    ws.row_dimensions[r].height = 16
    for ci, hdr in enumerate(["الشهر","الدرجة الكلية","التقييم","عدد المؤشرات","",""],1):
        sc(ws.cell(r,ci,hdr), bold=True, sz=9, color=WHITE,
           bg=DARK, ah="center", brd="inner")
    r += 1

    annual_scores = []
    for mi, (men, mar, msh) in enumerate(zip(MONTHS_EN, MONTHS_AR, MONTHS_SHORT)):
        ws.row_dimensions[r].height = 15
        rbg = LGRAY if mi % 2 == 0 else WHITE

        if emp_data is not None and not emp_data.empty:
            m_rows = emp_data[emp_data["Month"] == men]
        else:
            m_rows = pd.DataFrame()

        if not m_rows.empty:
            score = round(float(m_rows["KPI_%"].sum()), 1)
            n_kpi = len(m_rows)
            verb  = verbal(score)
            clr   = GREEN if score >= 80 else (YELLOW if score >= 60 else RED)
            annual_scores.append(score)
        else:
            score, verb, n_kpi, clr = "—", "—", "—", rbg

        sc(ws.cell(r,1,mar),          bold=True, sz=9, bg=rbg,   ah="center", brd="inner")
        sc(ws.cell(r,2,score),        bold=True, sz=9, bg=clr,   ah="center", brd="inner")
        sc(ws.cell(r,3,verb),         bold=True, sz=9, bg=clr,   ah="center", brd="inner")
        sc(ws.cell(r,4,n_kpi),        bold=False,sz=9, bg=rbg,   ah="center", brd="inner")
        r += 1

    # متوسط سنوي
    avg = round(sum(annual_scores)/len(annual_scores),1) if annual_scores else 0
    avg_clr = GREEN if avg >= 80 else (YELLOW if avg >= 60 else (RED if avg > 0 else LGRAY))
    ws.row_dimensions[r].height = 16
    mc(r,1,r,1, "المتوسط السنوي", bold=True, sz=9, bg=ORANGE,
       color=WHITE, ah="center", brd="outer")
    sc(ws.cell(r,2,f"{avg}%"),  bold=True, sz=10, color=DARK, bg=avg_clr, ah="center", brd="outer")
    sc(ws.cell(r,3,verbal(avg)),bold=True, sz=10, color=DARK, bg=avg_clr, ah="center", brd="outer")
    r += 2

    # ── تفاصيل المؤشرات لكل شهر (اختياري) ─────────────────────────
    if inc_monthly and emp_data is not None and not emp_data.empty:
        ws.row_dimensions[r].height = 18
        mc(r,1,r,6, "📈 تفاصيل مؤشرات الأداء الشهرية",
           bold=True, sz=11, color=WHITE, bg=DARK, ah="right", brd="outer")
        r += 1

        done_months = emp_data["Month"].dropna().unique().tolist()
        for men in MONTHS_EN:
            if men not in done_months:
                continue
            mar = MONTHS_AR[MONTHS_EN.index(men)]
            m_rows = emp_data[emp_data["Month"] == men]
            if m_rows.empty:
                continue

            # عنوان الشهر
            ws.row_dimensions[r].height = 16
            mc(r,1,r,6, f"▸ {mar}",
               bold=True, sz=10, color=WHITE, bg=MID, ah="right", brd="inner")
            r += 1

            # رأس جدول المؤشرات
            ws.row_dimensions[r].height = 14
            for ci,hdr in enumerate(["المؤشر","الوزن %","الدرجة","من 100","التقييم",""],1):
                sc(ws.cell(r,ci,hdr), bold=True, sz=8, color=WHITE,
                   bg=DARK, ah="center", brd="inner")
            r += 1

            for ki, (_, krow) in enumerate(m_rows.iterrows()):
                ws.row_dimensions[r].height = 14
                rbg = LGRAY if ki % 2 == 0 else WHITE
                kname  = str(krow.get("KPI_Name","")).strip()
                weight = float(krow.get("Weight",0) or 0)
                grade  = float(krow.get("KPI_%",0)  or 0)
                pct100 = round(grade/weight*100,1) if weight else 0
                vlbl   = verbal(pct100)
                kclr   = GREEN if pct100>=80 else (YELLOW if pct100>=60 else (RED if pct100>0 else rbg))

                sc(ws.cell(r,1,kname),  bold=False,sz=8, bg=rbg,  ah="right", wrap=True, brd="inner")
                sc(ws.cell(r,2,weight), bold=False,sz=8, bg=rbg,  ah="center",brd="inner")
                sc(ws.cell(r,3,grade),  bold=False,sz=8, bg=rbg,  ah="center",brd="inner")
                sc(ws.cell(r,4,pct100), bold=True, sz=8, bg=kclr, ah="center",brd="inner")
                sc(ws.cell(r,5,vlbl),   bold=True, sz=8, bg=kclr, ah="center",brd="inner")
                r += 1
            r += 1

    # ── إعداد الصفحة ────────────────────────────────────────────────
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4,right=0.4,top=0.5,bottom=0.5)
    ws.print_options.horizontalCentered = True

    # ── عمود A ذكي ──────────────────────────────────────────────────
    max_len = 0
    for row_cells in ws.iter_rows(min_col=1, max_col=1):
        for cell in row_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions["A"].width = min(max(max_len * 0.6 + 2, 28), 60)

    return ws


def render_cv_reports(df_emp, df_data, df_kpi, app_cfg, logo_path):
    try:
        import openpyxl
    except Exception as e:
        st.error(f"openpyxl مشكلة: {e}")
        return

    profiles = load_profiles()

    st.markdown("#### 📋 ملفات الموظفين")

    all_n = sorted(set(
        [str(n).strip() for n in df_emp["EmployeeName"].dropna()
         if str(n).strip() not in ("","nan")] + list(profiles.keys())
    ))

    if not all_n:
        st.info("لا يوجد موظفون.")
        return

    fc1, fc2 = st.columns(2)
    with fc1:
        mode = st.radio("نوع التقرير",
                        ["موظف محدد","قسم محدد","جميع الموظفين"],
                        horizontal=True)
    with fc2:
        inc = st.checkbox("تضمين التقييمات الشهرية التفصيلية", value=True)

    sel_year = st.selectbox("📅 السنة", [2025, 2026, 2027], key="cv_year")

    if mode == "موظف محدد":
        sel  = st.selectbox("اختر الموظف", all_n)
        tgts = [sel]
    elif mode == "قسم محدد":
        dpts = df_emp.iloc[:,2].dropna().astype(str).str.strip().unique().tolist()
        sd   = st.selectbox("القسم", sorted(dpts))
        tgts = df_emp[df_emp.iloc[:,2].astype(str).str.strip()==sd]["EmployeeName"].dropna().tolist()
    else:
        tgts = all_n

    if st.button("📥 إنشاء Excel", use_container_width=True):
        with st.spinner("جاري بناء التقارير..."):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for name in tgts:
                _build_cv_sheet(wb, name, df_emp, df_data, df_kpi,
                                profiles, inc, sel_year, app_cfg)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

        st.success(f"✅ تم إنشاء تقارير {len(tgts)} موظف")
        st.download_button(
            "⬇️ تحميل Excel",
            data=buf,
            file_name=f"ملفات_الموظفين_{date.today().strftime('%Y_%m_%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
