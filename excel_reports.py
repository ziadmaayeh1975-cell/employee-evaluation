import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


# ─── colour palette ───────────────────────────────────────────────────────────
DARK    = "1F3864"
MID     = "2E75B6"
ORANGE  = "ED7D31"
LGRAY   = "F2F2F2"
WHITE   = "FFFFFF"
YELLOW  = "FFF2CC"
GREEN_BG= "E2EFDA"
RED_BG  = "FFDAD9"
WARM    = "FFF3E0"
NOTE_BG = "FFFDE7"
TRAIN_BG= "F3E5F5"
INFO_BG = "EBF3FB"

_med = Side(style="medium", color="000000")
_thn = Side(style="thin",   color="000000")
BK   = Border(left=_med, right=_med, top=_med, bottom=_med)
TN   = Border(left=_thn, right=_thn, top=_thn, bottom=_thn)

_MAR = {"Jan":"يناير","Feb":"فبراير","Mar":"مارس","Apr":"أبريل",
        "May":"مايو",  "Jun":"يونيو", "Jul":"يوليو","Aug":"أغسطس",
        "Sep":"سبتمبر","Oct":"أكتوبر","Nov":"نوفمبر","Dec":"ديسمبر"}

MONTHS_LIST = ["يناير","فبراير","مارس","أبريل","مايو","يونيو",
               "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]

PERSONAL_KPIS = {
    "الالتزام بساعات الدوام اليومي وبمكان العمل",
    "الاهتمام بالمظهر العام والمداومة على المحافظة على علاقات إنسانية",
    "حاضر دائمًا ومتفانٍ ومحافظ بمحافظ بمحافظ ولا يُشار بمحافظ بمحافظ",
    "يتحمل ضغط العمل ولا يتذمر عند طلب أداء أعمال إضافية",
    "متحلٍ بالأمانة والمصداقية ولا يُفشي أسرار العمل أو الزملاء",
}

def verbal_grade(pct):
    if pct >= 90: return "ممتاز"
    if pct >= 80: return "جيد جداً"
    if pct >= 70: return "جيد"
    if pct >= 60: return "مقبول"
    return "ضعيف"

def kpi_score_to_pct(score, weight):
    return (score / weight * 100) if weight else 0

def rating_label(pct):
    return verbal_grade(pct)


def build_employee_sheet(wb, emp_name, job_title, dept, manager, year, kpis,
                         monthly_scores, notes="", training="",
                         chart_img=None, disciplinary_actions=None,
                         employee_id="", attendance_data=None):

    safe = emp_name[:28]
    if safe in [s.title for s in wb.worksheets]:
        safe = safe[:25] + "_2"
    ws = wb.create_sheet(safe)
    ws.sheet_view.rightToLeft  = True
    ws.sheet_view.showGridLines = False

    # ── helpers ──────────────────────────────────────────────────────────────
    def sc(cell, val=None, bold=False, sz=9, color="000000",
           bg=None, ah="right", av="center", wrap=False, brd=None):
        if val is not None:
            cell.value = val
        cell.font      = Font(name="Arial", bold=bold, size=sz, color=color)
        cell.alignment = Alignment(horizontal=ah, vertical=av,
                                   wrapText=wrap, readingOrder=2)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if brd:
            cell.border = brd

    def mc(r1, c1, r2, c2, val=None, **kw):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2, end_column=c2)
        sc(ws.cell(r1, c1, val), **kw)

    # ── pre-process monthly data ──────────────────────────────────────────────
    m_score, m_date, m_note, m_train = {}, {}, {}, {}
    for item in monthly_scores:
        ms = item[1]
        def _v(x): return str(x).strip() if x not in (None,"nan","None") else ""
        m_score[ms] = item[2]
        m_date[ms]  = _v(item[3]) if len(item) > 3 else ""
        m_note[ms]  = _v(item[4]) if len(item) > 4 else ""
        m_train[ms] = _v(item[5]) if len(item) > 5 else ""

    done = [(n,m,s) for n,m,s,*_ in monthly_scores if s > 0]
    pct  = sum(s for _,_,s in done) / len(done) * 100 if done else 0
    verb = verbal_grade(pct)
    sc_c = "375623" if pct >= 80 else ("C00000" if pct < 60 else "7F6000")
    sbg  = GREEN_BG if pct >= 80 else (YELLOW if pct >= 60 else RED_BG)

    job_kpis = [(k["KPI_Name"],k["Weight"],k.get("avg_score",0))
                for k in kpis if k["KPI_Name"] not in PERSONAL_KPIS]
    per_kpis = [(k["KPI_Name"],k["Weight"],k.get("avg_score",0))
                for k in kpis if k["KPI_Name"] in PERSONAL_KPIS]

    # company info
    _company, _branch = "مجموعة شركات فنون", ""
    try:
        from auth import load_app_settings as _las
        _cfg = _las()
        _company = _cfg.get("company_name", _company)
        _branch  = _cfg.get("branch_name",  "")
    except:
        pass
    _header = (f"نموذج تقييم الأداء السنوي — {_company}"
               + (f" — {_branch}" if _branch else ""))

    # ── disciplinary / attendance lookup ─────────────────────────────────────
    disc_by_month = {}
    disc_list = []  # for full disciplinary table (separate section)
    if disciplinary_actions is not None and not getattr(disciplinary_actions,"empty",True):
        for _, row_d in disciplinary_actions.iterrows():
            dd = row_d.get("action_date","")
            action_type = row_d.get("warning_type","")
            deduction = row_d.get("deduction_days",0)
            if dd:
                try:
                    mn = int(str(dd).split("-")[1])
                    disc_by_month.setdefault(mn, []).append(action_type)
                    disc_list.append({
                        "date": dd,
                        "type": action_type,
                        "reason": row_d.get("reason",""),
                        "deduction": deduction
                    })
                except: pass

    att_by_month = {}
    total_late_count = 0
    total_late_hours = 0.0
    if attendance_data is not None:
        if hasattr(attendance_data,"iterrows"):
            for _, row_a in attendance_data.iterrows():
                mn = row_a.get("month")
                late_count = row_a.get("late_count",0)
                late_hours = row_a.get("late_hours",0)
                if mn:
                    att_by_month[mn] = late_count
                    total_late_count += late_count
                    total_late_hours += late_hours
        elif isinstance(attendance_data, dict):
            mn = attendance_data.get("month")
            if mn:
                att_by_month[mn] = attendance_data.get("late_count",0)
                total_late_count = attendance_data.get("late_count",0)
                total_late_hours = attendance_data.get("late_hours",0)

    # ════════════════════════════════════════════════════════════════════════
    # COLUMN WIDTHS
    # Left block  A-E  (employee info + KPIs)
    # Divider     F
    # Right block G-N  (monthly table)
    # ════════════════════════════════════════════════════════════════════════
    col_w = {"A":5,"B":26,"C":16,"D":14,"E":3,
             "F":2,
             "G":10,"H":10,"I":14,"J":15,"K":24,"L":14,"M":14,"N":5}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # ROW 1 – full-width header
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 30
    mc(1,1,1,14, _header, bold=True, sz=11, color="FFFFFF", bg=DARK, ah="center")

    try:
        from openpyxl.drawing.image import Image as XLImage
        _logo = globals().get("LOGO_PATH","logo.png")
        if os.path.exists(_logo):
            img = XLImage(_logo)
            img.height, img.width = 60, 48
            img.anchor = "A1"
            ws.add_image(img)
    except: pass

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 2-8 – employee info (columns A-E)  +  monthly table header (G-N)
    # ════════════════════════════════════════════════════════════════════════
    INFO = [
        ("اسم الموظف",   emp_name),
        ("رقم الموظف",   employee_id),
        ("الوظيفة",      job_title),
        ("القسم",        dept),
        ("السنة",        str(year)),
        ("اسم المقيم",   manager),
        ("تاريخ التقييم",date.today().strftime("%d/%m/%Y")),
    ]

    # Monthly table title spans rows 2-3 on G-N
    ws.row_dimensions[2].height = 16
    mc(2,7,2,13, "نتيجة التقييم الشهري", bold=True, sz=10,
       color="FFFFFF", bg=DARK, ah="center")

    # Monthly table column headers on row 3
    ws.row_dimensions[3].height = 16
    mth_headers = ["الشهر","الدرجة (%)","التقييم اللفظي",
                   "تاريخ التقييم","ملاحظات المقيم","الإجراءات","عدد مرات التأخير"]
    for ci, h in enumerate(mth_headers, 7):
        sc(ws.cell(3, ci, h), bold=True, sz=8, color="FFFFFF", bg=MID, ah="center")

    # Employee info rows 2-8 (use columns A-E, skip D for spacing)
    for i, (lbl, val) in enumerate(INFO):
        row = 2 + i
        ws.row_dimensions[row].height = 16
        sc(ws.cell(row, 1, lbl),  bold=True, color="FFFFFF", bg=DARK, ah="center")
        sc(ws.cell(row, 2, val),  color="000000", bg=INFO_BG, ah="right")
        # Clear column D (spacer) if needed
        sc(ws.cell(row, 4, ""), bg=WHITE)

    # ════════════════════════════════════════════════════════════════════════
    # ROW 9 – annual result (left A-E) + first month data (right G-N)
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[9].height = 18
    mc(9,1,9,2, "نتيجة التقييم السنوي", bold=True, color="FFFFFF", bg=ORANGE, ah="center")
    mc(9,3,9,4, f"{int(round(pct))}% — {verb}",
       bold=True, sz=10, color=sc_c, bg=sbg, ah="center")

    # ════════════════════════════════════════════════════════════════════════
    # ROWS 4-15  monthly rows (right side G-N) start from row 4
    # ════════════════════════════════════════════════════════════════════════
    mth_start_row = 4
    for month_idx, month_name in enumerate(MONTHS_LIST, 1):
        mr = mth_start_row + (month_idx - 1)   # rows 4-15
        ws.row_dimensions[mr].height = 16
        rbg = LGRAY if month_idx % 2 == 0 else WHITE

        # find score
        month_data = None
        for item in monthly_scores:
            short  = item[1]
            mn_num = (list(_MAR.keys()).index(short)+1) if short in _MAR else 0
            if mn_num == month_idx:
                month_data = item; break

        if month_data and month_data[2] > 0:
            score      = month_data[2]
            eval_date  = month_data[3] if len(month_data) > 3 else ""
            note       = month_data[4] if len(month_data) > 4 else ""
            score_pct  = f"{round(score*100,1)}%"
            verbal_val = verbal_grade(score*100)
        else:
            score_pct = verbal_val = eval_date = note = "—"

        disc_text  = "، ".join(set(disc_by_month[month_idx])) \
                     if month_idx in disc_by_month else "—"
        late_count = att_by_month.get(month_idx, 0)
        late_txt   = str(late_count) if late_count > 0 else "—"

        sc(ws.cell(mr,  7, month_name),  bg=rbg, ah="center")
        sc(ws.cell(mr,  8, score_pct),   bg=rbg, ah="center")
        sc(ws.cell(mr,  9, verbal_val),  bg=rbg, ah="center")
        sc(ws.cell(mr, 10, str(eval_date) if eval_date else "—"), bg=rbg, ah="center")
        sc(ws.cell(mr, 11, note),        bg=rbg, ah="right", wrap=True)
        sc(ws.cell(mr, 12, disc_text),   bg=rbg, ah="center")
        sc(ws.cell(mr, 13, late_txt),    bg=rbg, ah="center")

    # ════════════════════════════════════════════════════════════════════════
    # KPI TABLES (left side, starting row 10)
    # ════════════════════════════════════════════════════════════════════════
    r = 10   # start KPI section right after annual result

    # ── Job KPIs header ──
    ws.row_dimensions[r].height = 16
    sc(ws.cell(r,1,"مؤشرات الأداء الوظيفي"), bold=True, color="FFFFFF", bg=DARK, ah="right")
    sc(ws.cell(r,2,"الوزن %"),               bold=True, color="FFFFFF", bg=DARK, ah="center")
    sc(ws.cell(r,3,"الدرجة (0-100)"),        bold=True, color="FFFFFF", bg=DARK, ah="center")
    sc(ws.cell(r,4,"التقييم"),               bold=True, color="FFFFFF", bg=DARK, ah="center")
    r += 1

    job_total_score, job_total_weight = 0.0, 0.0
    for i,(kname,weight,grade) in enumerate(job_kpis):
        rbg = LGRAY if i%2==0 else WHITE
        w, g = float(weight), float(grade)
        pct_val = round(kpi_score_to_pct(g,w),1)
        lbl     = rating_label(pct_val)
        job_total_score  += g
        job_total_weight += w
        kbg = (GREEN_BG if pct_val>=80 else (YELLOW if pct_val>=60
               else (RED_BG if pct_val>0 else rbg)))
        ws.row_dimensions[r].height = 15
        sc(ws.cell(r,1,kname),      bg=rbg, wrap=True, sz=8)
        sc(ws.cell(r,2,f"{w:.1f}%"),bg=rbg, ah="center", sz=8)
        sc(ws.cell(r,3,pct_val),    bold=True, bg=kbg, ah="center", sz=8)
        sc(ws.cell(r,4,lbl),        bold=True, bg=kbg, ah="center", sz=8)
        r += 1

    ws.row_dimensions[r].height = 15
    sc(ws.cell(r,1,"مجموع الأداء الوظيفي"),     bold=True, color="FFFFFF", bg=MID, ah="right")
    sc(ws.cell(r,2,f"{job_total_weight:.1f}%"),  bold=True, color="FFFFFF", bg=MID, ah="center")
    job_pct_total = round(kpi_score_to_pct(job_total_score,job_total_weight),1) \
                    if job_total_weight>0 else 0
    sc(ws.cell(r,3,f"{job_pct_total}%"), bold=True, color="FFFFFF", bg=MID, ah="center")
    sc(ws.cell(r,4,rating_label(job_pct_total)), bold=True, color="FFFFFF", bg=MID, ah="center")
    r += 2

    # ── Personal KPIs ──
    ws.row_dimensions[r].height = 15
    mc(r,1,r,4,"مؤشرات الصفات الشخصية",
       bold=True, color="FFFFFF", bg=ORANGE, ah="center")
    r += 1
    ws.row_dimensions[r].height = 15
    sc(ws.cell(r,1,"المؤشر"),         bold=True, color="FFFFFF", bg=MID, ah="right")
    sc(ws.cell(r,2,"الوزن %"),        bold=True, color="FFFFFF", bg=MID, ah="center")
    sc(ws.cell(r,3,"الدرجة (0-100)"), bold=True, color="FFFFFF", bg=MID, ah="center")
    sc(ws.cell(r,4,"التقييم"),        bold=True, color="FFFFFF", bg=MID, ah="center")
    r += 1

    per_total_score, per_total_weight = 0.0, 0.0
    for i,(kname,weight,grade) in enumerate(per_kpis):
        rbg = WARM if i%2==0 else WHITE
        w, g = float(weight), float(grade)
        pct_val = round(kpi_score_to_pct(g,w),1)
        lbl     = rating_label(pct_val)
        per_total_score  += g
        per_total_weight += w
        kbg = (GREEN_BG if pct_val>=80 else (YELLOW if pct_val>=60
               else (RED_BG if pct_val>0 else rbg)))
        ws.row_dimensions[r].height = 15
        sc(ws.cell(r,1,kname),      bg=rbg, wrap=True, sz=8)
        sc(ws.cell(r,2,f"{w:.1f}%"),bg=rbg, ah="center", sz=8)
        sc(ws.cell(r,3,pct_val),    bold=True, bg=kbg, ah="center", sz=8)
        sc(ws.cell(r,4,lbl),        bold=True, bg=kbg, ah="center", sz=8)
        r += 1

    ws.row_dimensions[r].height = 15
    sc(ws.cell(r,1,"مجموع الصفات الشخصية"),     bold=True, color="FFFFFF", bg=ORANGE, ah="right")
    sc(ws.cell(r,2,f"{per_total_weight:.1f}%"),  bold=True, color="FFFFFF", bg=ORANGE, ah="center")
    per_pct_total = round(kpi_score_to_pct(per_total_score,per_total_weight),1) \
                    if per_total_weight>0 else 0
    sc(ws.cell(r,3,f"{per_pct_total}%"),  bold=True, color="FFFFFF", bg=ORANGE, ah="center")
    sc(ws.cell(r,4,rating_label(per_pct_total)), bold=True, color="FFFFFF", bg=ORANGE, ah="center")
    r += 2

    # ════════════════════════════════════════════════════════════════════════
    # NEW SECTION: الإجراءات التأديبية المسجلة (Disciplinary Actions Table)
    # ════════════════════════════════════════════════════════════════════════
    if disc_list:
        ws.row_dimensions[r].height = 16
        mc(r,1,r,4, "الإجراءات التأديبية المسجلة",
           bold=True, color="FFFFFF", bg=DARK, ah="center")
        r += 1
        
        ws.row_dimensions[r].height = 15
        sc(ws.cell(r,1,"التاريخ"),    bold=True, color="FFFFFF", bg=MID, ah="center")
        sc(ws.cell(r,2,"نوع الإجراء"), bold=True, color="FFFFFF", bg=MID, ah="center")
        sc(ws.cell(r,3,"السبب"),      bold=True, color="FFFFFF", bg=MID, ah="center")
        sc(ws.cell(r,4,"خصم (أيام)"), bold=True, color="FFFFFF", bg=MID, ah="center")
        r += 1
        
        for i, act in enumerate(disc_list):
            rbg = LGRAY if i % 2 == 0 else WHITE
            ws.row_dimensions[r].height = 15
            sc(ws.cell(r,1,act["date"]),   bg=rbg, ah="center", sz=8)
            sc(ws.cell(r,2,act["type"]),   bg=rbg, ah="center", sz=8)
            sc(ws.cell(r,3,act["reason"]), bg=rbg, ah="right", wrap=True, sz=8)
            sc(ws.cell(r,4,act["deduction"]), bg=rbg, ah="center", sz=8)
            r += 1
        r += 1
    else:
        # Show empty table or skip
        ws.row_dimensions[r].height = 16
        mc(r,1,r,4, "الإجراءات التأديبية المسجلة",
           bold=True, color="FFFFFF", bg=DARK, ah="center")
        r += 1
        ws.row_dimensions[r].height = 15
        sc(ws.cell(r,1,"لا توجد إجراءات تأديبية مسجلة"), bold=False, bg=LGRAY, ah="center")
        sc(ws.cell(r,2,""), sc(ws.cell(r,3,"")), sc(ws.cell(r,4,"")))
        r += 2

    # ════════════════════════════════════════════════════════════════════════
    # NEW SECTION: الالتزام بالدوام (Attendance Summary)
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[r].height = 16
    mc(r,1,r,4, "الالتزام بالدوام",
       bold=True, color="FFFFFF", bg=DARK, ah="center")
    r += 1
    
    ws.row_dimensions[r].height = 15
    sc(ws.cell(r,1,"عدد مرات التأخير"), bold=True, color="FFFFFF", bg=MID, ah="center")
    sc(ws.cell(r,2,"إجمالي ساعات التأخير"), bold=True, color="FFFFFF", bg=MID, ah="center")
    sc(ws.cell(r,3,"ملاحظات"), bold=True, color="FFFFFF", bg=MID, ah="center")
    sc(ws.cell(r,4,""), bold=True, color="FFFFFF", bg=MID, ah="center")
    r += 1
    
    ws.row_dimensions[r].height = 15
    sc(ws.cell(r,1,str(total_late_count) if total_late_count > 0 else "0"), bg=LGRAY, ah="center", bold=True)
    sc(ws.cell(r,2,f"{total_late_hours:.2f}" if total_late_hours > 0 else "0"), bg=LGRAY, ah="center", bold=True)
    sc(ws.cell(r,3,""), bg=LGRAY, ah="right")
    sc(ws.cell(r,4,""), bg=LGRAY)
    r += 2

    # ── Notes / Training ──
    ws.row_dimensions[r].height = 20
    mc(r,1,r,4, f"ملاحظات المقيم: {notes or ''}", bg=NOTE_BG, wrap=True)
    r += 1

    _train_vals = [v for v in m_train.values()
                   if v and str(v).strip() not in ("","nan","None","—")]
    _train = _train_vals[0] if _train_vals else (training or "")
    mc(r,1,r,4,
       f"الاحتياجات التدريبية: {_train}" if _train else "الاحتياجات التدريبية:",
       bg=TRAIN_BG, wrap=True)
    r += 2

    # ── Signature ──
    ws.row_dimensions[r].height = 16
    sc(ws.cell(r,1,f"المسؤول المباشر: {manager}"), bold=True, ah="center")
    sc(ws.cell(r,2,"اسم الموظف"),                   bold=True, ah="center")
    sc(ws.cell(r,3,emp_name),                        bold=True, ah="right")
    r += 1
    ws.row_dimensions[r].height = 16
    sc(ws.cell(r,1,"التوقيع: _______________"), bold=True, bg=LGRAY, ah="center", brd=BK)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    sc(ws.cell(r,2,"التوقيع: _______________"), bold=True, bg=LGRAY, ah="center", brd=BK)

    # ── Page Setup ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)
    ws.print_options.horizontalCentered = True

    return ws


# ════════════════════════════════════════════════════════════════════════════
# DEMO – generates a sample file so you can verify the layout
# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    kpis = [
        {"KPI_Name":"جودة وقتاً الفحص واكتشاف الأعطال في كل حركات الخام وتكليل حالات الأرباح",
         "Weight":10,"avg_score":8},
        {"KPI_Name":"مطابقة نتائج الفحص لمتطلبات الجهات المعتمدة عند تقديم خدمة لتحسين المركبة",
         "Weight":9,"avg_score":8},
        {"KPI_Name":"نسبة تطبيق الأعمال من أول مرة أو بالإلزام بالأوقات المعيارية والإصلاح",
         "Weight":9,"avg_score":8},
        {"KPI_Name":"الإلتزام بتطبيقات نظام السلامة (HSE) واستخدام الأدوات من المخطط لمخالفات",
         "Weight":7,"avg_score":8},
        {"KPI_Name":"دقة وجودة التقارير الفنية واعتمادها من الجهة ذات التأثير أو الفائدة للتسهيلات",
         "Weight":8,"avg_score":8},
        {"KPI_Name":"دقة التوثيق وتسجيل بيانات الإجراءات ووثائق الإيجابات ذوي تأثير أو فائدة لبيانات",
         "Weight":8,"avg_score":8},
        {"KPI_Name":"سرعة الاستجابة العالية وتكليل من بعد ما عند الطلب",
         "Weight":6,"avg_score":8},
        {"KPI_Name":"سرعة التنسيق والتواصل الكامل لكليل التأخير الشتغيلي",
         "Weight":8,"avg_score":8},
        {"KPI_Name":"اختيار الاتتكات الداخلية بملاحظات مرة وسرعة معالجة الملاحظات",
         "Weight":6,"avg_score":8},
        {"KPI_Name":"نظافة وترتيب مكان العمل والأدوات والأجزاء ووفق المعايير المبنية",
         "Weight":9,"avg_score":8},
    ]

    personal_kpi_names = [
        "الالتزام بساعات الدوام اليومي وبمكان العمل",
        "الاهتمام بالمظهر العام والمداومة على المحافظة على علاقات إنسانية",
        "حاضر دائمًا ومتفانٍ ومحافظ بمحافظ بمحافظ ولا يُشار بمحافظ بمحافظ",
        "يتحمل ضغط العمل ولا يتذمر عند طلب أداء أعمال إضافية",
        "متحلٍ بالأمانة والمصداقية ولا يُفشي أسرار العمل أو الزملاء",
    ]
    for name in personal_kpi_names:
        kpis.append({"KPI_Name": name, "Weight": 4, "avg_score": 3.2})

    monthly = [
        ("بكر هشام سعيد حرب","Jan",0),
        ("بكر هشام سعيد حرب","Feb",0),
        ("بكر هشام سعيد حرب","Mar",0),
        ("بكر هشام سعيد حرب","Apr",0.8002,"29/04/2026","nan","nan"),
        ("بكر هشام سعيد حرب","May",0),
        ("بكر هشام سعيد حرب","Jun",0),
        ("بكر هشام سعيد حرب","Jul",0),
        ("بكر هشام سعيد حرب","Aug",0),
        ("بكر هشام سعيد حرب","Sep",0),
        ("بكر هشام سعيد حرب","Oct",0),
        ("بكر هشام سعيد حرب","Nov",0),
        ("بكر هشام سعيد حرب","Dec",0),
    ]

    # Sample disciplinary actions data (as DataFrame or list)
    import pandas as pd
    disciplinary_df = pd.DataFrame([
        {"action_date": "2026-04-07", "warning_type": "إداري أول", 
         "reason": "عدم التقييد بالتعليمات المعقدة", "deduction_days": 0}
    ])
    
    # Sample attendance data
    attendance_df = pd.DataFrame([
        {"month": 4, "late_count": 1, "late_hours": 0.35}
    ])

    build_employee_sheet(
        wb,
        emp_name        ="بكر هشام سعيد حرب",
        job_title       ="فني قسم محركات-كهربائي",
        dept            ="قسم الإرواع",
        manager         ="أب الفوائر",
        year            =2026,
        kpis            =kpis,
        monthly_scores  =monthly,
        notes           ="",
        training        ="",
        employee_id     ="397",
        disciplinary_actions=disciplinary_df,
        attendance_data=attendance_df,
    )

    out = "/mnt/user-data/outputs/employee_report_fixed.xlsx"
    wb.save(out)
    print(f"Saved → {out}")
