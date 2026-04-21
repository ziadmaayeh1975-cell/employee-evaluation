import pandas as pd
import streamlit as st
from constants import MONTHS_AR, MONTH_MAP, PERSONAL_KPIS, PERSONAL_WEIGHT
from calculations import verbal_grade, grade_color_hex, kpi_score_to_pct, calc_kpi_score, rating_label
from auth import get_current_reviewer, get_current_role

try:
    from database_manager import update_evaluation_in_db, delete_evaluation_from_db, db_exists
    _DB_OK = True
except ImportError:
    _DB_OK = False

from constants import FILE_PATH
def _can_edit_delete():
    """هل للمستخدم صلاحية التعديل والحذف؟ super_admin دائماً نعم."""
    from auth import get_current_role
    return get_current_role() in ("super_admin", "admin", "user")


def _reviewer_emp_set(df_emp):
    """
    None  = كل الموظفين (super_admin)
    set() = موظفو المقيم فقط
    """
    from auth import get_current_reviewer, get_current_role
    role             = get_current_role()
    current_reviewer = get_current_reviewer()

    # الأدمن الرئيسي → كل الموظفين دائماً
    if role == "super_admin":
        return None

    # أدمن عادي بدون reviewer → كل الموظفين
    if role == "admin" and not current_reviewer:
        return None

    # أدمن عادي أو user مع reviewer → موظفوه فقط
    reviewer_col = df_emp.columns[3] if len(df_emp.columns) > 3 else df_emp.columns[-1]
    return set(
        str(e).strip() for e in
        df_emp[df_emp[reviewer_col].astype(str).str.strip() == current_reviewer
               ]["EmployeeName"].dropna().tolist()
        if str(e).strip() not in ("","nan")
    )


def _reviewer_emp_list(df_emp):
    """يُعيد قائمة الموظفين المسموح برؤيتهم."""
    allowed = _reviewer_emp_set(df_emp)
    if allowed is None:
        return df_emp["EmployeeName"].dropna().astype(str).str.strip().tolist()
    return list(allowed)



def _safe_df(df):
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame(columns=[
            "EmployeeName","Month","KPI_Name","Weight","KPI_%",
            "Evaluator","Notes","Year","EvalDate","Training"
        ])
    df = df.copy()
    for col in ["EmployeeName","Month","KPI_Name","Weight","KPI_%","Year"]:
        if col not in df.columns:
            df[col] = pd.Series(dtype="object")
    return df


def _do_update(act_emp, act_month_en, act_year, new_grades):
    if _DB_OK and db_exists():
        return update_evaluation_in_db(act_emp, act_month_en, act_year, new_grades)
    import openpyxl
    wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True)
    ws = wb["DATA"]
    updated = 0
    for r in range(2, ws.max_row + 1):
        r_emp = str(ws.cell(r,1).value or "").strip()
        r_mon = str(ws.cell(r,2).value or "").strip()
        r_kpi = str(ws.cell(r,3).value or "").strip()
        try: r_yr = int(ws.cell(r,8).value or 2025)
        except: r_yr = 2025
        if r_emp==act_emp and r_mon==act_month_en and r_yr==int(act_year) and r_kpi in new_grades:
            ws.cell(r,5).value = float(new_grades[r_kpi]); updated += 1
    wb.save(FILE_PATH)
    return updated


def _do_delete(act_emp, act_month_en, act_year, kpi_to_del=None):
    if _DB_OK and db_exists():
        return delete_evaluation_from_db(act_emp, act_month_en, act_year, kpi_to_del)
    import openpyxl
    wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True)
    ws = wb["DATA"]
    rows_del = []
    for r in range(2, ws.max_row + 1):
        r_emp = str(ws.cell(r,1).value or "").strip()
        r_mon = str(ws.cell(r,2).value or "").strip()
        r_kpi = str(ws.cell(r,3).value or "").strip()
        try: r_yr = int(ws.cell(r,8).value or 2025)
        except: r_yr = 2025
        if r_emp==act_emp and r_mon==act_month_en and r_yr==int(act_year):
            if kpi_to_del is None or r_kpi==kpi_to_del:
                rows_del.append(r)
    for r in sorted(rows_del, reverse=True):
        ws.delete_rows(r)
    wb.save(FILE_PATH)
    return len(rows_del)


def render_manage(df_emp, df_kpi, df_data):
    st.subheader("👁️ عرض تقييم الموظف – تعديل وحذف التقييم الشهري")

    df_data = _safe_df(df_data)

    if df_data.empty or df_data["EmployeeName"].dropna().empty:
        st.info("لا توجد بيانات تقييم بعد.")
        return

    # ── فلترة حسب المقيم ────────────────────────────────────────
    allowed_emps = _reviewer_emp_set(df_emp)

    mg1, mg2, mg3 = st.columns(3)
    with mg1:
        all_emps = sorted([
            str(e) for e in df_data["EmployeeName"].dropna().unique()
            if str(e).strip() not in ("","nan")
            and (allowed_emps is None or str(e).strip() in allowed_emps)
        ])
        mg_emps = ["-- الكل --"] + all_emps
        mg_emp  = st.selectbox("👤 الموظف", mg_emps, key="mg_emp")
    with mg2:
        mg_month = st.selectbox("📅 الشهر", ["-- الكل --"] + MONTHS_AR, key="mg_month")
    with mg3:
        mg_year = st.selectbox("🗓️ السنة", ["-- الكل --", 2025, 2026, 2027], key="mg_year")

    view_df = df_data.copy()
    # تطبيق فلتر المقيم
    if allowed_emps is not None:
        view_df = view_df[view_df["EmployeeName"].isin(allowed_emps)]
    if mg_emp   != "-- الكل --": view_df = view_df[view_df["EmployeeName"] == mg_emp]
    if mg_month != "-- الكل --": view_df = view_df[view_df["Month"] == MONTH_MAP.get(mg_month, mg_month)]
    if mg_year  != "-- الكل --": view_df = view_df[view_df["Year"] == int(mg_year)]

    if view_df.empty:
        st.warning("لا توجد سجلات تطابق هذا البحث.")
        return

    summary_view = (
        view_df.groupby(["EmployeeName","Month","Year"])
        .agg(مجموع_الدرجات=("KPI_%","sum"), عدد_المؤشرات=("KPI_%","count"))
        .reset_index()
        .rename(columns={"EmployeeName":"الموظف","Month":"الشهر","Year":"السنة"})
    )
    summary_view["التقييم %"] = summary_view["مجموع_الدرجات"].apply(lambda x: f"{int(round(x))}%")
    summary_view["نتيجة"]     = summary_view["مجموع_الدرجات"].apply(verbal_grade)

    st.markdown("#### 📋 السجلات الموجودة")
    st.dataframe(summary_view[["الموظف","الشهر","السنة","التقييم %","نتيجة","عدد_المؤشرات"]],
                 hide_index=True, use_container_width=True)

    st.write("---")
    st.markdown("#### 🔧 اختر سجل للتعديل أو الحذف")
    mg4, mg5, mg6 = st.columns(3)
    with mg4:
        act_emps = sorted([
            str(e) for e in view_df["EmployeeName"].dropna().unique()
            if str(e).strip() not in ("","nan")
        ])
        act_emp = st.selectbox("👤 الموظف", act_emps, key="act_emp")
    with mg5:
        act_months_en = [
            str(m) for m in view_df[view_df["EmployeeName"]==act_emp]["Month"].dropna().unique()
            if str(m).strip() not in ("","nan")
        ]
        act_month_en = st.selectbox("📅 الشهر", act_months_en, key="act_month")
    with mg6:
        act_years = sorted([
            int(y) for y in view_df[
                (view_df["EmployeeName"]==act_emp) & (view_df["Month"]==act_month_en)
            ]["Year"].dropna().unique()
        ])
        act_year = st.selectbox("🗓️ السنة", act_years, key="act_year")

    sel_records = view_df[
        (view_df["EmployeeName"]==act_emp) &
        (view_df["Month"]==act_month_en) &
        (view_df["Year"]==act_year)
    ].copy()

    if sel_records.empty:
        st.warning("لا توجد سجلات لهذا الاختيار.")
        return

    total_pct = sel_records["KPI_%"].sum()
    st.markdown(f"""
    <div style="background:#F8FAFC;border:1px solid #CBD5E1;border-radius:10px;
                padding:14px;margin:10px 0;direction:rtl;">
        <b>👤 {act_emp}</b> | <b>📅 {act_month_en}</b> | <b>🗓️ {act_year}</b>
        &nbsp;&nbsp;
        <span style="color:{grade_color_hex(total_pct)};font-size:1.2rem;font-weight:bold;">
            {total_pct:.1f}% – {verbal_grade(total_pct)}
        </span>
    </div>""", unsafe_allow_html=True)

    job_recs  = sel_records[~sel_records["KPI_Name"].isin(PERSONAL_KPIS)]
    pers_recs = sel_records[sel_records["KPI_Name"].isin(PERSONAL_KPIS)]

    if not job_recs.empty:
        st.markdown("**🎯 مؤشرات الأداء الوظيفي:**")
        jdf = pd.DataFrame([{
            "المؤشر":           row["KPI_Name"],
            "الوزن النسبي (%)": row["Weight"],
            "الدرجة (0-100)":   round(kpi_score_to_pct(float(row["KPI_%"]), float(row["Weight"])), 1),
            "التقييم":          rating_label(kpi_score_to_pct(float(row["KPI_%"]), float(row["Weight"])))
        } for _, row in job_recs.iterrows()])
        st.dataframe(jdf, hide_index=True, use_container_width=True)

    if not pers_recs.empty:
        st.markdown(f"**🌟 مؤشرات الصفات الشخصية** — المجموع: **{pers_recs['KPI_%'].sum():.1f}%**")
        pdfx = pd.DataFrame([{
            "المؤشر":           row["KPI_Name"],
            "الوزن النسبي (%)": row["Weight"],
            "الدرجة (0-100)":   round(kpi_score_to_pct(float(row["KPI_%"]), float(row["Weight"])), 1),
            "التقييم":          rating_label(kpi_score_to_pct(float(row["KPI_%"]), float(row["Weight"])))
        } for _, row in pers_recs.iterrows()])
        st.dataframe(pdfx, hide_index=True, use_container_width=True)

    st.write("---")
    col_edit, col_del = st.columns(2)

    with col_edit:
        st.markdown("##### ✏️ تعديل الدرجات")
        if not job_recs.empty:
            st.markdown("🎯 **مؤشرات الأداء الوظيفي**")
            with st.form("edit_form_job"):
                new_grades = {}
                for _, row in job_recs.iterrows():
                    cur_pct = round(kpi_score_to_pct(float(row["KPI_%"]), float(row["Weight"])))
                    st.markdown(f"<div style='font-size:12px;font-weight:700;color:#1E3A8A;margin-bottom:2px;'>{row['KPI_Name']}</div>", unsafe_allow_html=True)
                    new_val = st.number_input("", min_value=0, max_value=100,
                        value=int(cur_pct), step=1,
                        key=f"edit_job_{row['KPI_Name']}", label_visibility="collapsed")
                    new_grades[row["KPI_Name"]] = calc_kpi_score(float(new_val), float(row["Weight"]))
                st.info(f"المجموع: **{sum(new_grades.values())}%**")
                if st.form_submit_button("💾 حفظ تعديلات مؤشرات الأداء", use_container_width=True):
                    try:
                        u = _do_update(act_emp, act_month_en, act_year, new_grades)
                        st.success(f"✅ تم تعديل {u} مؤشر!")
                        st.cache_data.clear(); st.rerun()
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")

        if not pers_recs.empty:
            st.markdown("🌟 **مؤشرات الصفات الشخصية**")
            with st.form("edit_form_personal"):
                new_pers = {}
                for _, row in pers_recs.iterrows():
                    pval = st.number_input(f"{row['KPI_Name']} (الوزن: {row['Weight']}%)",
                        min_value=0, max_value=int(row["Weight"]), value=int(row["KPI_%"]),
                        step=1, key=f"edit_pers_{row['KPI_Name']}")
                    new_pers[row["KPI_Name"]] = pval
                st.info(f"المجموع: **{sum(new_pers.values())}%**")
                if st.form_submit_button("💾 حفظ تعديلات الصفات الشخصية", use_container_width=True):
                    try:
                        u = _do_update(act_emp, act_month_en, act_year, new_pers)
                        st.success(f"✅ تم تعديل {u} مؤشر!")
                        st.cache_data.clear(); st.rerun()
                    except Exception as e:
                        st.error(f"❌ خطأ: {e}")

    with col_del:
        st.markdown("##### 🗑️ حذف التقييم")
        st.warning(f"سيتم حذف **{len(sel_records)} مؤشر** لـ **{act_emp}** في **{act_month_en} {act_year}**.")
        del_option = st.radio("ماذا تريد أن تحذف؟", ["حذف التقييم كامل","حذف مؤشر محدد فقط"], key="del_option")
        kpi_to_del = None
        if del_option == "حذف مؤشر محدد فقط":
            kpi_to_del = st.selectbox("اختر المؤشر", sel_records["KPI_Name"].tolist(), key="kpi_del_sel")
        confirm = st.checkbox("✅ أؤكد رغبتي في الحذف", key="del_confirm")
        if st.button("🗑️ تنفيذ الحذف", disabled=not confirm, use_container_width=True, type="primary"):
            try:
                deleted = _do_delete(act_emp, act_month_en, act_year, kpi_to_del)
                st.success(f"✅ تم حذف {deleted} سجل!")
                st.cache_data.clear(); st.rerun()
            except Exception as e:
                st.error(f"❌ خطأ: {e}")
