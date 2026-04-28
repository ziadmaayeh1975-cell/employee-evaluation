"""
data_loader.py — يقرأ من قاعدة البيانات أولاً، Excel احتياطياً
مع دعم الأعمدة ديناميكياً في شيتي EMPLOYEES و KPIs
"""
import pandas as pd
import streamlit as st
from datetime import date
from constants import FILE_PATH, MONTH_MAP

try:
    from database_manager import load_data_from_db, save_evaluation_to_db, db_exists
    _DB_AVAILABLE = True
except ImportError:
    _DB_AVAILABLE = False


def _normalize_col(col: str) -> str:
    """توحيد اسم العمود: إزالة المسافات وتحويل للأحرف الصغيرة"""
    return str(col).strip().lower().replace(" ", "").replace("_", "")


def _map_employee_columns(df_emp):
    """إعادة تسمية أعمدة EMPLOYEES ديناميكياً بغض النظر عن حجم الحروف"""
    mapping = {}
    for col in df_emp.columns:
        c = _normalize_col(col)
        if any(x in c for x in ["رقمالموظف", "employeeid", "empid"]):
            mapping[col] = "رقم الموظف"
        elif any(x in c for x in ["employeename", "اسمالموظف"]):
            mapping[col] = "EmployeeName"
        elif any(x in c for x in ["jobtitle", "الوظيفة", "وظيفة"]):
            mapping[col] = "JobTitle"
        elif any(x in c for x in ["قسم", "department", "dept"]):
            mapping[col] = "القسم"
        elif any(x in c for x in ["مقيم", "evaluator", "reviewer"]):
            mapping[col] = "اسم المقيم"
    return df_emp.rename(columns=mapping)


def _map_kpi_columns(df_kpi):
    """إعادة تسمية أعمدة KPIs ديناميكياً بغض النظر عن حجم الحروف"""
    mapping = {}
    for col in df_kpi.columns:
        c = _normalize_col(col)
        if any(x in c for x in ["jobtitle", "الوظيفة", "وظيفة"]):
            mapping[col] = "JobTitle"
        elif any(x in c for x in ["kpiname", "اسمالمؤشر", "المؤشر"]):
            mapping[col] = "KPI_Name"
        elif any(x in c for x in ["weight", "الوزن", "وزن"]):
            mapping[col] = "Weight"
    return df_kpi.rename(columns=mapping)


def _map_data_columns(df_data):
    """إعادة تسمية أعمدة DATA ديناميكياً بغض النظر عن حجم الحروف"""
    mapping = {}
    for col in df_data.columns:
        c = _normalize_col(col)
        if any(x in c for x in ["employeename", "اسمالموظف"]):
            mapping[col] = "EmployeeName"
        elif any(x in c for x in ["month", "الشهر", "شهر"]):
            mapping[col] = "Month"
        elif any(x in c for x in ["kpiname", "اسمالمؤشر", "المؤشر"]):
            mapping[col] = "KPI_Name"
        elif any(x in c for x in ["kpi%", "kpipct", "kpipercent", "نسبة", "درجة"]):
            mapping[col] = "KPI_%"
        elif any(x in c for x in ["weight", "الوزن", "وزن"]):
            mapping[col] = "Weight"
        elif any(x in c for x in ["evaluator", "مقيم", "reviewer"]):
            mapping[col] = "Evaluator"
        elif any(x in c for x in ["notes", "nots", "ملاحظات", "ملاحظة"]):
            mapping[col] = "Notes"
        elif any(x in c for x in ["year", "السنة", "سنة"]):
            mapping[col] = "Year"
        elif any(x in c for x in ["evaldate", "entrydate", "تاريخ"]):
            mapping[col] = "EvalDate"
        elif any(x in c for x in ["training", "تدريب", "احتياجات"]):
            mapping[col] = "Training"
    return df_data.rename(columns=mapping)


@st.cache_data(ttl=30)
def load_data():
    if _DB_AVAILABLE and db_exists():
        return load_data_from_db()

    try:
        import openpyxl
        df_emp  = pd.read_excel(FILE_PATH, sheet_name="EMPLOYEES")
        df_kpi  = pd.read_excel(FILE_PATH, sheet_name="KPIs")
        df_data = pd.read_excel(FILE_PATH, sheet_name="DATA")

        # ── تنظيف أسماء الأعمدة وقيم النصوص لكل الشيتات ──
        for df in [df_emp, df_kpi, df_data]:
            if not df.empty:
                df.columns = [str(c).strip() for c in df.columns]
                for col in df.select_dtypes("object").columns:
                    df[col] = df[col].astype(str).str.strip()

        # ── إعادة تسمية الأعمدة (case-insensitive) ──
        df_emp  = _map_employee_columns(df_emp)
        df_kpi  = _map_kpi_columns(df_kpi)
        df_data = _map_data_columns(df_data)

        # ── التحقق من أعمدة EMPLOYEES ──
        required_emp_cols = ["رقم الموظف", "EmployeeName", "JobTitle", "القسم", "اسم المقيم"]
        missing_emp = [col for col in required_emp_cols if col not in df_emp.columns]
        if missing_emp:
            st.error(f"⚠️ الأعمدة المطلوبة غير موجودة في EMPLOYEES: {missing_emp}")
            return None, None, None

        # ── التحقق من أعمدة KPIs ──
        required_kpi_cols = ["JobTitle", "KPI_Name", "Weight"]
        missing_kpi = [col for col in required_kpi_cols if col not in df_kpi.columns]
        if missing_kpi:
            st.error(f"⚠️ الأعمدة المطلوبة غير موجودة في KPIs: {missing_kpi}")
            return None, None, None

        # ── التحقق من أعمدة DATA ──
        required_data_cols = ["EmployeeName", "Month", "KPI_Name", "KPI_%"]
        missing_data = [col for col in required_data_cols if col not in df_data.columns]
        if missing_data:
            st.error(f"⚠️ الأعمدة المطلوبة غير موجودة في DATA: {missing_data}")
            return None, None, None

        # ── ترتيب الأعمدة ──
        df_emp = df_emp[required_emp_cols]
        df_kpi = df_kpi[required_kpi_cols]

        # ── معالجة عمود Year ──
        if "Year" not in df_data.columns:
            df_data["Year"] = 2025
        else:
            df_data["Year"] = pd.to_numeric(df_data["Year"], errors="coerce").fillna(2025).astype(int)

        # ── إضافة الأعمدة الاختيارية إن لم تكن موجودة ──
        for col in ["EvalDate", "Training", "Notes", "Evaluator"]:
            if col not in df_data.columns:
                df_data[col] = ""

        # ── التأكد من أن KPI_% رقمية ──
        df_data["KPI_%"] = pd.to_numeric(df_data["KPI_%"], errors="coerce").fillna(0.0)

        # ── التأكد من أن Weight رقمية في KPIs ──
        df_kpi["Weight"] = pd.to_numeric(df_kpi["Weight"], errors="coerce").fillna(0.0)

        return df_emp, df_kpi, df_data

    except Exception as e:
        st.error(f"❌ تعذّر تحميل البيانات: {e}")
        return None, None, None


def save_evaluation(emp_name, month_ar, year, manager, dept,
                    kpi_rows, notes="", training=""):
    if _DB_AVAILABLE and db_exists():
        ok, err = save_evaluation_to_db(emp_name, month_ar, year, manager, dept,
                                        kpi_rows, notes, training)
        if ok:
            load_data.clear()
        return ok, err

    try:
        import openpyxl
        wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True)
        ws = wb["DATA"]

        if ws.cell(1, 8).value != "Year":
            ws.cell(1, 8).value = "Year"
        if ws.cell(1, 9).value != "EvalDate":
            ws.cell(1, 9).value = "EvalDate"
        if ws.cell(1, 10).value != "Training":
            ws.cell(1, 10).value = "Training"

        month_en   = MONTH_MAP.get(month_ar, month_ar)
        eval_date  = date.today().strftime("%d/%m/%Y")
        nr         = ws.max_row + 1

        for item in kpi_rows:
            if len(item) == 4:
                kpi_name, weight, grade, rating_lbl = item
            else:
                kpi_name, weight, grade = item[:3]
                rating_lbl = ""

            ws.cell(nr, 1).value = emp_name
            ws.cell(nr, 2).value = month_en
            ws.cell(nr, 3).value = kpi_name
            ws.cell(nr, 4).value = float(weight)
            ws.cell(nr, 5).value = round(float(grade), 2)
            ws.cell(nr, 6).value = manager
            ws.cell(nr, 7).value = notes
            ws.cell(nr, 8).value = int(year)
            ws.cell(nr, 9).value = eval_date
            ws.cell(nr, 10).value = training
            ws.cell(nr, 11).value = rating_lbl
            nr += 1

        wb.save(FILE_PATH)
        return True, None
    except Exception as e:
        return False, str(e)


def get_emp_notes(emp_name):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(FILE_PATH, read_only=True, keep_vba=True)
        ws = wb["INPUT"]
        notes    = ws["B20"].value or ""
        training = ws["B21"].value or ""
        wb.close()
        return str(notes), str(training)
    except:
        return "", ""
