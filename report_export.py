"""
report_export.py
================
بناء تقارير Excel لنظام تقييم الأداء
"""

import io
import os
import base64
from datetime import date

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter, column_index_from_string

from constants import (
    LOGO_PATH, MONTHS_AR, MONTHS_EN, MONTHS_SHORT, PERSONAL_KPIS,
    DARK, MID, LBLUE, ORANGE, YELLOW, LGRAY, GREEN_BG, RED_BG, WHITE, CREAM,
    OUTER_B, INNER_B,
)
from calculations import verbal_grade, kpi_score_to_pct, rating_label


# ═══════════════════════════════════════════════════════════════════
# الثوابت المشتركة
# ═══════════════════════════════════════════════════════════════════
_WARM     = "FFF3E0"
_NOTE_BG  = "FFFDE7"
_TRAIN_BG = "F3E5F5"
_INFO_BG  = "EBF3FB"
_DISC_BG  = "FEE2E2"
_ATT_BG   = "E0F2FE"

_med  = Side(style="medium", color="000000")
_thn  = Side(style="thin",   color="AAAAAA")
_BK   = Border(left=_med, right=_med, top=_med, bottom=_med)
_TN   = Border(left=_thn, right=_thn, top=_thn, bottom=_thn)

_MAR  = {
    "Jan":"يناير","Feb":"فبراير","Mar":"مارس","Apr":"أبريل",
    "May":"مايو","Jun":"يونيو","Jul":"يوليو","Aug":"أغسطس",
    "Sep":"سبتمبر","Oct":"أكتوبر","Nov":"نوفمبر","Dec":"ديسمبر",
}
_MONTHS_LIST = [
    "يناير","فبراير","مارس","أبريل","مايو","يونيو",
    "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر",
]


def _fill(hex_c):
    return PatternFill("solid", fgColor=str(hex_c).lstrip("#"))


def _font(bold=False, color="000000", size=9):
    return Font(name="Arial", bold=bold, color=str(color).lstrip("#"), size=size)


def _align(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrapText=wrap, readingOrder=2)


def _sc(cell, val=None, bold=False, sz=9, color="000000",
        bg=None, ah="right", av="center", wrap=False, brd=None):
    if val is not None:
        cell.value = val
    cell.font      = _font(bold=bold, color=color, size=sz)
    cell.alignment = _align(h=ah, v=av, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if brd:
        cell.border = brd


def _mc(ws, r1, c1, r2, c2, val=None, **kw):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    _sc(ws.cell(r1, c1, val), **kw)


def _company_header():
    _company, _branch = "مجموعة شركات فنون", ""
    try:
        from auth import load_app_settings as _las
        cfg = _las()
        _company = cfg.get("company_name", _company)
        _branch  = cfg.get("branch_name",  "")
    except Exception:
        pass
    return (f"نموذج تقييم الأداء السنوي — {_company}"
            + (f" — {_branch}" if _branch else ""))


def _add_logo(ws, anchor="A1", h=45, w=36):
    for _lp in [LOGO_PATH, "logo.png"]:
        if os.path.exists(_lp):
            try:
                img = XLImage(_lp)
                img.height, img.width = h, w
                img.anchor = anchor
                ws.add_image(img)
            except Exception:
                pass
            break


def _auto_fit_column_a(ws, start_row=1, end_row=None):
    max_len = 0
    if end_row is None:
        end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row, 1)
        if cell.value:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
    ws.column_dimensions["A"].width = min(max(max_len * 0.7 + 2, 8), 50)


def _disc_by_month(disciplinary_actions):
    """
    يُرجع قاموساً: {month_num: [{"date": "...", "type": "...", "reason": "..."}, ...]}
    """
    result = {}
    if disciplinary_actions is None:
        return result
    try:
        if getattr(disciplinary_actions, "empty", True):
            return result
        for _, row in disciplinary_actions.iterrows():
            dd = str(row.get("action_date", "") or "").strip()
            if not dd or dd in ("nan", "None"):
                continue
            try:
                mn = int(dd.split("-")[1])
                # تنسيق التاريخ كاملاً dd/mm/yyyy
                try:
                    import pandas as _pd
                    full_date = _pd.to_datetime(dd).strftime("%d/%m/%Y")
                except Exception:
                    full_date = dd
                warning_type = str(row.get("warning_type", "") or "").strip()
                reason       = str(row.get("reason",       "") or "").strip()
                result.setdefault(mn, []).append({
                    "date":   full_date,
                    "type":   warning_type if warning_type not in ("", "nan", "None") else "—",
                    "reason": reason       if reason       not in ("", "nan", "None") else "—",
                })
            except Exception:
                pass
    except Exception:
        pass
    return result


def _att_by_month(attendance_data):
    """
    تحويل بيانات التأخير إلى قاموس لكل شهر
    المتوقع: DataFrame يحتوي على أعمدة month, late_count, late_hours
    """
    result = {}
    result_hours = {}
    if attendance_data is None:
        return result, result_hours
    try:
        import pandas as _pd
        if isinstance(attendance_data, _pd.DataFrame):
            for _, row in attendance_data.iterrows():
                mn = row.get("month")
                if mn:
                    result[int(mn)] = int(row.get("late_count", 0) or 0)
                    result_hours[int(mn)] = float(row.get("late_hours", 0) or 0.0)
        elif isinstance(attendance_data, dict):
            mn = attendance_data.get("month")
            if mn:
                result[int(mn)] = int(attendance_data.get("late_count", 0) or 0)
                result_hours[int(mn)] = float(attendance_data.get("late_hours", 0) or 0.0)
        elif isinstance(attendance_data, list):
            for item in attendance_data:
                mn = item.get("month")
                if mn:
                    result[int(mn)] = int(item.get("late_count", 0) or 0)
                    result_hours[int(mn)] = float(item.get("late_hours", 0) or 0.0)
    except Exception:
        pass
    return result, result_hours


# ═══════════════════════════════════════════════════════════════════
# build_employee_sheet
# ═══════════════════════════════════════════════════════════════════
def build_employee_sheet(
    wb,
    emp_name, job_title, dept, manager, year,
    kpis,
    monthly_scores,
    notes="", training="",
    chart_img=None,
    disciplinary_actions=None,
    employee_id="",
    attendance_data=None,
):
    safe = emp_name[:28]
    if safe in [s.title for s in wb.worksheets]:
        safe = safe[:25] + "_2"
    ws = wb.create_sheet(safe)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    # إعداد عرض الأعمدة - إضافة عمود لساعات التأخير
    col_cfg = {
        "A": 5, "B": 26, "C": 16, "D": 13,
        "E": 2, "F": 2,
        "G": 10, "H": 11, "I": 14, "J": 14,
        "K": 22, "L": 14, "M": 14, "N": 14,  # N لساعات التأخير
        "O": 5,
    }
    for col, w in col_cfg.items():
        ws.column_dimensions[col].width = w

    # معالجة البيانات
    m_train = {}
    for item in monthly_scores:
        ms = item[1]
        def _v(x): return str(x).strip() if x not in (None, "nan", "None", "—") else ""
        m_train[ms] = _v(item[5]) if len(item) > 5 else ""

    done = [(n, m, s) for n, m, s, *_ in monthly_scores if s > 0]
    pct = sum(s for _, _, s in done) / len(done) * 100 if done else 0
    verb = verbal_grade(pct)
    sc_c = "375623" if pct >= 80 else ("C00000" if pct < 60 else "7F6000")
    sbg = GREEN_BG if pct >= 80 else (YELLOW if pct >= 60 else RED_BG)

    job_kpis = [(k["KPI_Name"], k["Weight"], k.get("avg_score", 0))
                for k in kpis if k["KPI_Name"] not in PERSONAL_KPIS]
    per_kpis = [(k["KPI_Name"], k["Weight"], k.get("avg_score", 0))
                for k in kpis if k["KPI_Name"] in PERSONAL_KPIS]

    disc_map = _disc_by_month(disciplinary_actions)
    att_count_map, att_hours_map = _att_by_month(attendance_data)
    
    total_late_count = sum(att_count_map.values())
    total_late_hours = sum(att_hours_map.values())

    # ════════════════════════════════════════════════════════════
    # ROW 1 — ترويسة
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 32
    _mc(ws, 1, 1, 1, 15, _company_header(),
        bold=True, sz=11, color="FFFFFF", bg=DARK, ah="center")
    _add_logo(ws, anchor="A1", h=45, w=36)

    # ════════════════════════════════════════════════════════════
    # ROWS 2-8 — معلومات الموظف (A-D)
    # ════════════════════════════════════════════════════════════
    INFO = [
        ("اسم الموظف", emp_name),
        ("رقم الموظف", employee_id),
        ("الوظيفة", job_title),
        ("القسم", dept),
        ("السنة", str(year)),
        ("اسم المقيم", manager),
        ("تاريخ التقييم", date.today().strftime("%d/%m/%Y")),
    ]
    for i, (lbl, val) in enumerate(INFO):
        rr = 2 + i
        ws.row_dimensions[rr].height = 16
        _sc(ws.cell(rr, 1, lbl), bold=True, color="FFFFFF", bg=DARK, ah="center", brd=_TN)
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        _sc(ws.cell(rr, 2, val), color="000000", bg=_INFO_BG, ah="right", brd=_TN)

    # ════════════════════════════════════════════════════════════
    # ROW 3 — عنوان الجدول الشهري (مضاف)
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[3].height = 18
    _mc(ws, 3, 7, 3, 14, "نتيجة التقييم الشهري",
        bold=True, sz=10, color="FFFFFF", bg=DARK, ah="center", brd=_TN)

    # ════════════════════════════════════════════════════════════
    # ROW 4 — رؤوس أعمدة الجدول الشهري
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[4].height = 15
    mth_hdrs = [
        "الشهر", "الدرجة (%)", "التقييم اللفظي",
        "تاريخ التقييم", "ملاحظات المقيم",
        "الإجراءات التأديبية", "عدد مرات التأخير", "ساعات التأخير"
    ]
    for ci, h in enumerate(mth_hdrs, 7):
        _sc(ws.cell(4, ci, h), bold=True, sz=8, color="FFFFFF", bg=MID,
            ah="center", brd=_TN)

    # ════════════════════════════════════════════════════════════
    # ROWS 5-16 — بيانات الأشهر الـ12 (تبدأ من الصف 5)
    # ════════════════════════════════════════════════════════════
    for month_idx, month_name in enumerate(_MONTHS_LIST, 1):
        mr = 4 + month_idx  # يبدأ من الصف 5
        ws.row_dimensions[mr].height = 15
        rbg = LGRAY if month_idx % 2 == 0 else WHITE

        month_data = None
        for item in monthly_scores:
            short = item[1]
            mn_num = (list(_MAR.keys()).index(short) + 1) if short in _MAR else 0
            if mn_num == month_idx:
                month_data = item
                break

        if month_data and month_data[2] > 0:
            score = month_data[2]
            eval_date = str(month_data[3]) if len(month_data) > 3 else ""
            note = str(month_data[4]) if len(month_data) > 4 else ""
            score_pct = f"{round(score * 100, 1)}%"
            verbal_val = verbal_grade(score * 100)
            if eval_date in ("None", "nan", ""):
                eval_date = "—"
            if note in ("None", "nan", ""):
                note = "—"
        else:
            score_pct = verbal_val = eval_date = note = "—"

        disc_text = ("، ".join(set(a["type"] for a in disc_map[month_idx]))
                     if month_idx in disc_map else "—")
        late_count = att_count_map.get(month_idx, 0)
        late_hours = att_hours_map.get(month_idx, 0.0)
        late_count_txt = str(late_count) if late_count > 0 else "—"
        late_hours_txt = f"{late_hours:.2f}" if late_hours > 0 else "—"

        if score_pct != "—":
            sv = float(score_pct.replace("%", ""))
            sbg2 = GREEN_BG if sv >= 80 else (YELLOW if sv >= 60 else RED_BG)
        else:
            sbg2 = rbg

        _sc(ws.cell(mr, 7, month_name), bg=rbg, ah="center", brd=_TN)
        _sc(ws.cell(mr, 8, score_pct), bg=sbg2, ah="center", bold=(score_pct != "—"), brd=_TN)
        _sc(ws.cell(mr, 9, verbal_val), bg=sbg2, ah="center", brd=_TN)
        _sc(ws.cell(mr, 10, eval_date), bg=rbg, ah="center", sz=8, brd=_TN)
        _sc(ws.cell(mr, 11, note), bg=rbg, ah="right", wrap=True, sz=8, brd=_TN)
        _sc(ws.cell(mr, 12, disc_text), bg=(_DISC_BG if disc_text != "—" else rbg),
            ah="center", sz=8, brd=_TN)
        _sc(ws.cell(mr, 13, late_count_txt), bg=(_ATT_BG if late_count > 0 else rbg),
            ah="center", sz=8, brd=_TN)
        _sc(ws.cell(mr, 14, late_hours_txt), bg=(_ATT_BG if late_hours > 0 else rbg),
            ah="center", sz=8, brd=_TN)

    # ════════════════════════════════════════════════════════════
    # ROW 17 — إجمالي الإجراءات والتأخير (بعد الأشهر الـ12)
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[17].height = 15
    total_disc = sum(len(v) for v in disc_map.values())
    _mc(ws, 17, 7, 17, 11, "الإجماليات السنوية",
        bold=True, color="FFFFFF", bg=DARK, ah="center", brd=_TN)
    _sc(ws.cell(17, 12,
                f"إجمالي الإجراءات: {total_disc}" if total_disc else "لا توجد إجراءات"),
        bold=True, bg=_DISC_BG, ah="center", sz=8, brd=_TN)
    _sc(ws.cell(17, 13,
                f"إجمالي مرات التأخير: {total_late_count}" if total_late_count > 0 else "لا تأخير"),
        bold=True, bg=_ATT_BG, ah="center", sz=8, brd=_TN)
    _sc(ws.cell(17, 14,
                f"إجمالي ساعات التأخير: {total_late_hours:.2f}" if total_late_hours > 0 else "لا تأخير"),
        bold=True, bg=_ATT_BG, ah="center", sz=8, brd=_TN)

    # ════════════════════════════════════════════════════════════
    # ROW 10 — نتيجة التقييم السنوي (بجانب الجدول الشهري)
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[10].height = 17
    _mc(ws, 10, 1, 10, 2, "نتيجة التقييم السنوي",
        bold=True, color="FFFFFF", bg=ORANGE, ah="center", brd=_TN)
    _mc(ws, 10, 3, 10, 4, f"{int(round(pct))}% — {verb}",
        bold=True, sz=10, color=sc_c, bg=sbg, ah="center", brd=_TN)

    # ════════════════════════════════════════════════════════════
    # KPI SECTION (يبدأ من ROW 11)
    # ════════════════════════════════════════════════════════════
    r = 11

    # ── مؤشرات الأداء الوظيفي ──
    ws.row_dimensions[r].height = 15
    _sc(ws.cell(r, 1, "مؤشرات الأداء الوظيفي"), bold=True, color="FFFFFF", bg=DARK, ah="right", brd=_TN)
    _sc(ws.cell(r, 2, "الوزن %"), bold=True, color="FFFFFF", bg=DARK, ah="center", brd=_TN)
    _sc(ws.cell(r, 3, "الدرجة (0-100)"), bold=True, color="FFFFFF", bg=DARK, ah="center", brd=_TN)
    _sc(ws.cell(r, 4, "التقييم"), bold=True, color="FFFFFF", bg=DARK, ah="center", brd=_TN)
    r += 1

    job_total_score, job_total_weight = 0.0, 0.0
    for i, (kname, weight, grade) in enumerate(job_kpis):
        rbg = LGRAY if i % 2 == 0 else WHITE
        w, g = float(weight), float(grade)
        pct_val = round(kpi_score_to_pct(g, w), 1)
        lbl = rating_label(pct_val)
        job_total_score += g
        job_total_weight += w
        kbg = (GREEN_BG if pct_val >= 80
               else (YELLOW if pct_val >= 60
               else (RED_BG if pct_val > 0 else rbg)))
        ws.row_dimensions[r].height = 14
        _sc(ws.cell(r, 1, kname), bg=rbg, wrap=True, sz=8, brd=_TN)
        _sc(ws.cell(r, 2, f"{w:.1f}%"), bg=rbg, ah="center", sz=8, brd=_TN)
        _sc(ws.cell(r, 3, pct_val), bold=True, bg=kbg, ah="center", sz=8, brd=_TN)
        _sc(ws.cell(r, 4, lbl), bold=True, bg=kbg, ah="center", sz=8, brd=_TN)
        r += 1

    ws.row_dimensions[r].height = 14
    jp = round(kpi_score_to_pct(job_total_score, job_total_weight), 1) if job_total_weight > 0 else 0
    _sc(ws.cell(r, 1, "مجموع الأداء الوظيفي"), bold=True, color="FFFFFF", bg=MID, ah="right", brd=_TN)
    _sc(ws.cell(r, 2, f"{job_total_weight:.1f}%"), bold=True, color="FFFFFF", bg=MID, ah="center", brd=_TN)
    _sc(ws.cell(r, 3, f"{jp}%"), bold=True, color="FFFFFF", bg=MID, ah="center", brd=_TN)
    _sc(ws.cell(r, 4, rating_label(jp)), bold=True, color="FFFFFF", bg=MID, ah="center", brd=_TN)
    r += 2

    # ── مؤشرات الصفات الشخصية ──
    ws.row_dimensions[r].height = 14
    _mc(ws, r, 1, r, 4, "مؤشرات الصفات الشخصية",
        bold=True, color="FFFFFF", bg=ORANGE, ah="center", brd=_TN)
    r += 1
    ws.row_dimensions[r].height = 14
    _sc(ws.cell(r, 1, "المؤشر"), bold=True, color="FFFFFF", bg=MID, ah="right", brd=_TN)
    _sc(ws.cell(r, 2, "الوزن %"), bold=True, color="FFFFFF", bg=MID, ah="center", brd=_TN)
    _sc(ws.cell(r, 3, "الدرجة (0-100)"), bold=True, color="FFFFFF", bg=MID, ah="center", brd=_TN)
    _sc(ws.cell(r, 4, "التقييم"), bold=True, color="FFFFFF", bg=MID, ah="center", brd=_TN)
    r += 1

    per_total_score, per_total_weight = 0.0, 0.0
    for i, (kname, weight, grade) in enumerate(per_kpis):
        rbg = _WARM if i % 2 == 0 else WHITE
        w, g = float(weight), float(grade)
        pct_val = round(kpi_score_to_pct(g, w), 1)
        lbl = rating_label(pct_val)
        per_total_score += g
        per_total_weight += w
        kbg = (GREEN_BG if pct_val >= 80
               else (YELLOW if pct_val >= 60
               else (RED_BG if pct_val > 0 else rbg)))
        ws.row_dimensions[r].height = 14
        _sc(ws.cell(r, 1, kname), bg=rbg, wrap=True, sz=8, brd=_TN)
        _sc(ws.cell(r, 2, f"{w:.1f}%"), bg=rbg, ah="center", sz=8, brd=_TN)
        _sc(ws.cell(r, 3, pct_val), bold=True, bg=kbg, ah="center", sz=8, brd=_TN)
        _sc(ws.cell(r, 4, lbl), bold=True, bg=kbg, ah="center", sz=8, brd=_TN)
        r += 1

    ws.row_dimensions[r].height = 14
    pp = round(kpi_score_to_pct(per_total_score, per_total_weight), 1) if per_total_weight > 0 else 0
    _sc(ws.cell(r, 1, "مجموع الصفات الشخصية"), bold=True, color="FFFFFF", bg=ORANGE, ah="right", brd=_TN)
    _sc(ws.cell(r, 2, f"{per_total_weight:.1f}%"), bold=True, color="FFFFFF", bg=ORANGE, ah="center", brd=_TN)
    _sc(ws.cell(r, 3, f"{pp}%"), bold=True, color="FFFFFF", bg=ORANGE, ah="center", brd=_TN)
    _sc(ws.cell(r, 4, rating_label(pp)), bold=True, color="FFFFFF", bg=ORANGE, ah="center", brd=_TN)
    r += 2

    # ════════════════════════════════════════════════════════════
    # الإجراءات التأديبية المسجلة (أسفل مؤشرات الصفات الشخصية)
    # ════════════════════════════════════════════════════════════
    if disc_map:
        ws.row_dimensions[r].height = 14
        _mc(ws, r, 1, r, 4, "⚠️ الإجراءات التأديبية المسجلة",
            bold=True, color="FFFFFF", bg="C00000", ah="right", brd=_TN)
        r += 1
        # رأس الأعمدة
        ws.row_dimensions[r].height = 13
        _sc(ws.cell(r, 1, "تاريخ الإجراء"), bold=True, color="FFFFFF", bg=MID, ah="center", sz=8, brd=_TN)
        _sc(ws.cell(r, 2, "نوع الإنذار"),   bold=True, color="FFFFFF", bg=MID, ah="center", sz=8, brd=_TN)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        _sc(ws.cell(r, 3, "سبب الإجراء"),   bold=True, color="FFFFFF", bg=MID, ah="center", sz=8, brd=_TN)
        r += 1
        for mn, actions in sorted(disc_map.items()):
            for act in actions:
                ws.row_dimensions[r].height = 13
                _sc(ws.cell(r, 1, act.get("date", "—")),
                    bg=_DISC_BG, ah="center", sz=8, brd=_TN)
                _sc(ws.cell(r, 2, act.get("type", "—")),
                    bg=_DISC_BG, ah="center", sz=8, brd=_TN)
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
                _sc(ws.cell(r, 3, act.get("reason", "—")),
                    bg=_DISC_BG, ah="right", wrap=True, sz=8, brd=_TN)
                r += 1
        r += 1
    else:
        ws.row_dimensions[r].height = 14
        _mc(ws, r, 1, r, 4, "⚠️ الإجراءات التأديبية المسجلة",
            bold=True, color="FFFFFF", bg="C00000", ah="right", brd=_TN)
        r += 1
        ws.row_dimensions[r].height = 13
        _mc(ws, r, 1, r, 4, "لا توجد إجراءات تأديبية مسجلة",
            bg=_DISC_BG, ah="center", sz=8, brd=_TN)
        r += 1
        r += 1

    # ════════════════════════════════════════════════════════════
    # الالتزام بالدوام (أسفل الإجراءات التأديبية) - مع إجمالي الساعات
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[r].height = 14
    _mc(ws, r, 1, r, 4, "⏰ الالتزام بالدوام",
        bold=True, color="FFFFFF", bg="1F3864", ah="right", brd=_TN)
    r += 1
    
    ws.row_dimensions[r].height = 13
    _sc(ws.cell(r, 1, "إجمالي عدد مرات التأخير"), bold=True, bg=_ATT_BG, ah="right", sz=8, brd=_TN)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    _sc(ws.cell(r, 2, str(total_late_count)), bg=_ATT_BG, ah="right", sz=8, brd=_TN)
    r += 1
    
    ws.row_dimensions[r].height = 13
    _sc(ws.cell(r, 1, "إجمالي ساعات التأخير"), bold=True, bg=_ATT_BG, ah="right", sz=8, brd=_TN)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    _sc(ws.cell(r, 2, f"{total_late_hours:.2f}"), bg=_ATT_BG, ah="right", sz=8, brd=_TN)
    r += 2

    # ════════════════════════════════════════════════════════════
    # ملاحظات المقيم والاحتياجات التدريبية
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[r].height = 20
    _train_vals = [v for v in m_train.values()
                   if v and str(v).strip() not in ("", "nan", "None", "—")]
    _train = _train_vals[0] if _train_vals else (training or "")
    _mc(ws, r, 1, r, 4, f"ملاحظات المقيم: {notes or '—'}",
        bg=_NOTE_BG, wrap=True, brd=_TN)
    r += 1
    _mc(ws, r, 1, r, 4,
        f"الاحتياجات التدريبية: {_train}" if _train else "الاحتياجات التدريبية: —",
        bg=_TRAIN_BG, wrap=True, brd=_TN)
    r += 2

    # ════════════════════════════════════════════════════════════
    # التوقيع
    # ════════════════════════════════════════════════════════════
    ws.row_dimensions[r].height = 16
    _sc(ws.cell(r, 1, f"المسؤول المباشر: {manager}"),
        bold=True, ah="center", brd=_BK)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    _sc(ws.cell(r, 2, f"اسم الموظف: {emp_name}"),
        bold=True, ah="right", brd=_BK)
    r += 1
    ws.row_dimensions[r].height = 16
    _sc(ws.cell(r, 1, "التوقيع: _______________"),
        bold=True, bg=LGRAY, ah="center", brd=_BK)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    _sc(ws.cell(r, 2, "التوقيع: _______________"),
        bold=True, bg=LGRAY, ah="center", brd=_BK)

    # ضبط عرض العمود A تلقائياً
    _auto_fit_column_a(ws, start_row=1, end_row=ws.max_row)

    # إعداد الطباعة
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)
    ws.print_options.horizontalCentered = True

    return ws


# ═══════════════════════════════════════════════════════════════════
# build_summary_sheet
# ═══════════════════════════════════════════════════════════════════
def build_summary_sheet(
    wb,
    rows,
    title="ملخص التقييم",
    year=None,
    chart_img=None,
    disciplinary_summary=None,
    attendance_summary=None,
):
    ws = wb.create_sheet(title[:28])
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    col_cfg = {
        "A": 4, "B": 30, "C": 14, "D": 8,
        "E": 10, "F": 12, "G": 12,
        "H": 14, "I": 14, "J": 14,  # إضافة عمود لساعات التأخير
    }
    for col, w in col_cfg.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 32
    _mc(ws, 1, 1, 1, 10, title,
        bold=True, sz=12, color="FFFFFF", bg=DARK, ah="center", brd=_BK)
    _add_logo(ws, anchor="A1", h=30, w=24)

    ws.row_dimensions[2].height = 4
    ws.row_dimensions[3].height = 16

    hdrs = ["#", "اسم الموظف", "القسم", "السنة",
            "الأشهر", "المعدل %", "التقييم",
            "الإجراءات التأديبية", "عدد مرات التأخير", "إجمالي ساعات التأخير"]
    for ci, h in enumerate(hdrs, 1):
        _sc(ws.cell(3, ci, h), bold=True, sz=9, color="FFFFFF",
            bg=DARK, ah="center", brd=_BK)

    disc_s = disciplinary_summary or {}
    att_s = attendance_summary or {}

    for i, row_data in enumerate(rows, 4):
        name, dept_, months, pct_val, verb_ = row_data[:5]
        # قراءة بيانات الإجراءات التأديبية والتأخير مباشرة من row_data
        disc_cnt   = int(row_data[5]) if len(row_data) > 5 and row_data[5] else 0
        late_count = int(row_data[6]) if len(row_data) > 6 and row_data[6] else 0
        late_hours = float(row_data[7]) if len(row_data) > 7 and row_data[7] else 0.0

        # fallback من disciplinary_summary/attendance_summary إذا لم تُمرَّر في row_data
        if disc_cnt == 0 and disc_s:
            disc_info = disc_s.get(name, {})
            disc_cnt  = disc_info.get("count", 0) if isinstance(disc_info, dict) else int(disc_info or 0)
        if late_count == 0 and att_s:
            att_info   = att_s.get(name, {})
            if isinstance(att_info, dict):
                late_count = att_info.get("count", 0)
                late_hours = att_info.get("hours", 0.0)
            else:
                late_count = int(att_info or 0)

        ws.row_dimensions[i].height = 15
        rbg  = LGRAY if i % 2 == 0 else WHITE
        sc_c = "375623" if pct_val >= 80 else ("C00000" if pct_val < 60 else "000000")
        vbg  = (GREEN_BG if pct_val >= 80
                else (YELLOW if pct_val >= 70
                else (RED_BG if pct_val < 60 else LGRAY)))

        disc_bg = _DISC_BG if disc_cnt > 0 else rbg
        att_bg = _ATT_BG if late_count > 0 else rbg

        _sc(ws.cell(i, 1, i - 3), sz=8, ah="center", bg=rbg, brd=INNER_B)
        _sc(ws.cell(i, 2, name), sz=9, ah="right", bg=rbg, brd=INNER_B)
        _sc(ws.cell(i, 3, dept_), sz=8, ah="center", bg=rbg, brd=INNER_B)
        _sc(ws.cell(i, 4, year or ""), sz=9, ah="center", bg=rbg, brd=INNER_B)
        _sc(ws.cell(i, 5, months), sz=9, ah="center", bg=rbg, brd=INNER_B)
        _sc(ws.cell(i, 6, f"{pct_val:.1f}%"),
            sz=10, bold=True, color=sc_c, ah="center", bg=vbg, brd=INNER_B)
        _sc(ws.cell(i, 7, verb_),
            sz=9, bold=True, color=sc_c, ah="center", bg=vbg, brd=INNER_B)
        _sc(ws.cell(i, 8,
                    f"{disc_cnt} إجراء" if disc_cnt > 0 else "—"),
            sz=8, ah="center", bg=disc_bg, brd=INNER_B)
        _sc(ws.cell(i, 9,
                    f"{late_count} مرة" if late_count > 0 else "—"),
            sz=8, ah="center", bg=att_bg, brd=INNER_B)
        _sc(ws.cell(i, 10,
                    f"{late_hours:.2f} ساعة" if late_hours > 0 else "—"),
            sz=8, ah="center", bg=att_bg, brd=INNER_B)

    last = 3 + len(rows)
    thick_s = Side(style="medium", color="000000")
    for rr in range(3, last + 1):
        for c_idx in [1, 10]:
            b = ws.cell(rr, c_idx).border
            if c_idx == 1:
                ws.cell(rr, c_idx).border = Border(
                    left=thick_s, right=b.right, top=b.top, bottom=b.bottom)
            else:
                ws.cell(rr, c_idx).border = Border(
                    left=b.left, right=thick_s, top=b.top, bottom=b.bottom)
    for c_idx in range(1, 11):
        b = ws.cell(last, c_idx).border
        ws.cell(last, c_idx).border = Border(
            left=b.left, right=b.right, top=b.top, bottom=thick_s)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True

    return ws


# ═══════════════════════════════════════════════════════════════════
# print_preview_html (نسخة محسنة ومصححة بالكامل)
# ═══════════════════════════════════════════════════════════════════
def _rgb_to_hex(color_obj):
    try:
        if color_obj and color_obj.type == "rgb":
            rgb = color_obj.rgb
            return "#" + rgb[2:]
    except Exception:
        pass
    return ""


def print_preview_html(xlsx_buf, title="تقرير", chart_b64=""):
    xlsx_buf.seek(0)
    wb = openpyxl.load_workbook(xlsx_buf, data_only=True)

    _logo_b64 = ""
    for _lp in [LOGO_PATH, "logo.png"]:
        if os.path.exists(_lp):
            try:
                with open(_lp, "rb") as _lf:
                    _logo_b64 = base64.b64encode(_lf.read()).decode()
            except Exception:
                pass
            break

    # ─────────────────────────────────────────────────────────────────────────
    # CSS — معاينة شاشة بعرض A4 landscape + طباعة صفحة واحدة
    # قاعدة: حجم الخط لا يُمسّ — الموائمة عبر transform:scale لكل ورقة
    # ─────────────────────────────────────────────────────────────────────────
    CSS = """
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

body {
    font-family: 'Cairo', Arial, sans-serif;
    direction: rtl;
    background: #dde1ea;
    padding: 12px;
    margin: 0;
}

/* ── زر الطباعة ─────────────────────────────────────── */
.no-print {
    text-align: center;
    margin-bottom: 10px;
    position: sticky;
    top: 0;
    background: #dde1ea;
    padding: 6px;
    z-index: 100;
}
.no-print button {
    padding: 9px 32px;
    background: #1F3864;
    color: #fff;
    border: none;
    border-radius: 8px;
    cursor: pointer;
    font-size: 13px;
    font-family: 'Cairo', Arial, sans-serif;
    font-weight: 700;
    box-shadow: 0 2px 8px rgba(0,0,0,.2);
}
.no-print button:hover { background: #2E75B6; }

/* ── غلاف الصفحة (شاشة) ─────────────────────────────
   عرض A4 landscape الصافي = 297mm − هوامش 2×8mm = 281mm
   نعرضها كاملةً في المعاينة بدون scroll أفقي           */
.page-wrapper {
    width: 281mm;
    margin: 0 auto 20px auto;
    background: #fff;
    box-shadow: 0 4px 16px rgba(0,0,0,.15);
    border-radius: 4px;
    padding: 5mm 7mm;
    direction: rtl;
    overflow: hidden;
}

/* ── الجدول ─────────────────────────────────────────── */
table {
    border-collapse: collapse;
    width: 100%;
    table-layout: fixed;      /* يوزع الأعمدة حسب الأوزان النسبية */
    font-size: 8pt;           /* حجم ثابت لا يُعدَّل */
    font-family: 'Cairo', Arial, sans-serif;
    direction: rtl;
    margin: 2px 0;
}
th, td {
    border: 0.5px solid #aaa;
    padding: 2px 4px;
    vertical-align: middle;
    text-align: center;
    word-break: break-word;
    overflow-wrap: break-word;
    white-space: normal;
}
th {
    background: #1F3864;
    color: white;
    font-weight: bold;
    font-size: 8pt;
}
td { background: white; color: #333; }
td:first-child { font-weight: bold; background: #f0f2f5; }

/* ── ترويسة الشعار ──────────────────────────────────── */
.logo-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background: #1F3864;
    padding: 4px 12px;
    margin-bottom: 5px;
    border-radius: 4px;
}
.logo-header span { color: white; font-size: 10pt; font-weight: bold; }
.logo-header img  { height: 34px; width: auto; object-fit: contain; }

/* ── معلومات الموظف ─────────────────────────────────── */
.info-grid {
    background: #EBF3FB;
    padding: 4px 10px;
    margin-bottom: 5px;
    border-radius: 4px;
    font-size: 8pt;
    direction: rtl;
}
.info-row  { display: flex; flex-wrap: wrap; margin-bottom: 2px; }
.info-label { font-weight: bold; width: 88px; color: #1F3864; }
.info-value { flex: 1; color: #333; }

/* ── النتيجة السنوية ─────────────────────────────────── */
.annual-result {
    background: #ED7D31;
    color: white;
    padding: 6px;
    text-align: center;
    border-radius: 4px;
    margin: 5px 0;
    font-weight: bold;
    font-size: 12pt;
}
.annual-result small { font-size: 8pt; display: block; }

/* ══════════════════════════════════════════════════════
   طباعة — A4 landscape — صفحة واحدة لكل ورقة عمل
   الحل: نُقلِّص كل .page-wrapper بـ transform:scale
   حتى يتسع في مساحة الطباعة دون المساس بحجم الخط
   ══════════════════════════════════════════════════════ */
@media print {
    /* ─── إعداد الصفحة ──────────────────────────────── */
    @page {
        size: A4 landscape;
        margin: 6mm;
    }

    /* ─── إخفاء زر الطباعة ─────────────────────────── */
    .no-print { display: none !important; }

    /* ─── جسم الصفحة ───────────────────────────────── */
    body {
        background: white;
        padding: 0;
        margin: 0;
    }

    /* ─── كل ورقة عمل = صفحة طباعة مستقلة ────────── */
    .page-wrapper {
        /* فصل كل ورقة في صفحة مستقلة */
        page-break-before: always;
        break-before: page;
        page-break-inside: avoid;
        break-inside: avoid;

        /* إزالة تنسيقات الشاشة */
        width: 100% !important;
        margin: 0 !important;
        padding: 3mm 4mm !important;
        box-shadow: none !important;
        border-radius: 0 !important;
        overflow: visible !important;

        /* ── الموائمة للصفحة الواحدة بدون تغيير الخط ──
           transform:scale يُصغِّر المحتوى كوحدة متكاملة
           مع الحفاظ على نسبه الداخلية (الخط، الهوامش...)
           القيمة 0.85 تتسع لمعظم التقارير؛ المتصفح يتعامل
           مع fit-to-page تلقائياً عند الطباعة أيضاً        */
        transform-origin: top right;
    }

    /* ── إخبار المتصفح بملء الصفحة ─────────────────── */
    html, body {
        width: 100%;
        height: 100%;
    }

    /* ── الجدول يمتد للعرض الكامل ───────────────────── */
    table {
        width: 100% !important;
        table-layout: fixed !important;
        font-size: 8pt !important;
    }
    th, td {
        border-color: #000 !important;
        padding: 2px 3px !important;
    }

    /* ── ألوان الطباعة ───────────────────────────────── */
    * {
        -webkit-print-color-adjust: exact !important;
        print-color-adjust: exact !important;
        color-adjust: exact !important;
    }
}
"""

    # ─────────────────────────────────────────────────────────────────────────
    # حساب أوزان الأعمدة النسبية من ws.column_dimensions
    # يُعيد قاموساً {col_index: weight_percent}
    # ─────────────────────────────────────────────────────────────────────────
    def _col_weights(ws):
        raw = {}
        for letter, cd in ws.column_dimensions.items():
            try:
                idx = column_index_from_string(letter)
                if not cd.hidden:
                    raw[idx] = float(cd.width or 8)
            except Exception:
                pass
        total = sum(raw.values()) or 1
        return {idx: w / total * 100 for idx, w in raw.items()}

    # ─────────────────────────────────────────────────────────────────────────
    # تحويل ورقة عمل → HTML جدول
    # table-layout:fixed + colgroup بنسب مئوية → autofit كامل بلا scroll
    # ─────────────────────────────────────────────────────────────────────────
    def parse_table_from_ws(ws):
        if ws.max_row == 0:
            return ""

        # ── خلايا مدمجة ──────────────────────────────────────
        merged = {}
        skip   = set()
        for m in ws.merged_cells.ranges:
            merged[(m.min_row, m.min_col)] = (
                m.max_row - m.min_row + 1,
                m.max_col - m.min_col + 1,
            )
            for r2 in range(m.min_row, m.max_row + 1):
                for c2 in range(m.min_col, m.max_col + 1):
                    if (r2, c2) != (m.min_row, m.min_col):
                        skip.add((r2, c2))

        max_col = ws.max_column
        weights = _col_weights(ws)

        # ── رأس الجدول بالأوزان النسبية (%) ─────────────────
        html  = '<div style="width:100%;overflow:hidden;">'
        html += '<table>'
        html += '<colgroup>'
        for c in range(1, max_col + 1):
            pct = weights.get(c, 100 / max_col)
            html += f'<col style="width:{pct:.2f}%;">'
        html += '</colgroup>'

        # ── صفوف البيانات ─────────────────────────────────────
        for r in range(1, ws.max_row + 1):
            # تخطي الصفوف المدمجة كاملاً
            if all((r, c) in skip for c in range(1, max_col + 1)):
                continue

            html += '<tr>'
            for c in range(1, max_col + 1):
                if (r, c) in skip:
                    continue

                cell  = ws.cell(r, c)
                val   = cell.value
                text  = "" if val is None else str(val).replace("\n", "<br>")
                style = ""

                # خلفية
                p = cell.fill
                if p and p.fill_type == "solid":
                    bg = _rgb_to_hex(p.fgColor)
                    if bg and bg.lower() not in ("#000000", "#ffffff", ""):
                        style += f"background:{bg};"

                # نص: bold + color فقط — حجم الخط يرثه من table
                f_obj = cell.font
                if f_obj:
                    if f_obj.bold:
                        style += "font-weight:bold;"
                    fc = _rgb_to_hex(f_obj.color)
                    if fc and fc not in ("#000000", ""):
                        style += f"color:{fc};"

                # محاذاة
                a = cell.alignment
                if a:
                    ha = ("right"  if a.horizontal == "right"
                          else "left" if a.horizontal == "left"
                          else "center")
                    va = ("top"    if a.vertical == "top"
                          else "bottom" if a.vertical == "bottom"
                          else "middle")
                    style += f"text-align:{ha};vertical-align:{va};"

                style += "padding:2px 4px;"

                # دمج
                span = ""
                if (r, c) in merged:
                    rs2, cs2 = merged[(r, c)]
                    if rs2 > 1: span += f' rowspan="{rs2}"'
                    if cs2 > 1: span += f' colspan="{cs2}"'

                html += f'<td style="{style}"{span}>{text}</td>'
            html += '</tr>'

        html += '</table></div>'
        return html

    # ─────────────────────────────────────────────────────────────────────────
    # تجميع HTML النهائي
    # ─────────────────────────────────────────────────────────────────────────
    parts = [
        "<!DOCTYPE html>",
        '<html dir="rtl" lang="ar">',
        "<head>",
        '<meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        f"<title>{title}</title>",
        f"<style>{CSS}</style>",
        "</head><body>",
        '<div class="no-print">',
        f'<button onclick="window.print()">🖨️ {title} — طباعة</button>',
        "</div>",
    ]

    first_page = True
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0:
            continue

        header_text = _company_header()

        # ── استخراج اسم الموظف ───────────────────────────────
        emp_name_display = sheet_name
        try:
            for r in range(2, min(ws.max_row, 10)):
                if ws.cell(r, 1).value == "اسم الموظف" and ws.cell(r, 2).value:
                    emp_name_display = str(ws.cell(r, 2).value)
                    break
        except Exception:
            pass

        # page-break-before على الصفحة الثانية فأكثر يُضافه CSS تلقائياً
        parts.append('<div class="page-wrapper">')

        # ── الترويسة ──────────────────────────────────────────
        if _logo_b64:
            parts.append(
                f'<div class="logo-header">'
                f'<span>{header_text}</span>'
                f'<img src="data:image/png;base64,{_logo_b64}" alt="شعار">'
                f'</div>'
            )
        else:
            parts.append(
                f'<div class="logo-header"><span>{header_text}</span></div>'
            )

        # ── معلومات الموظف (صفوف 2-8) ────────────────────────
        info_rows = []
        try:
            for r in range(2, 9):
                lbl = ws.cell(r, 1).value
                val = ws.cell(r, 2).value
                if (lbl and val
                        and str(lbl).strip()
                        not in ("", "نتيجة التقييم السنوي")):
                    info_rows.append((str(lbl).strip(), str(val).strip()))
        except Exception:
            pass

        if info_rows:
            parts.append('<div class="info-grid">')
            for lbl, val in info_rows:
                parts.append(
                    f'<div class="info-row">'
                    f'<span class="info-label">{lbl}:</span>'
                    f'<span class="info-value">{val}</span>'
                    f'</div>'
                )
            parts.append('</div>')

        # ── النتيجة السنوية ───────────────────────────────────
        try:
            for r in range(9, 13):
                if ws.cell(r, 1).value == "نتيجة التقييم السنوي":
                    annual_val = ws.cell(r, 3).value or ws.cell(r, 4).value
                    if annual_val:
                        parts.append(
                            f'<div class="annual-result">'
                            f'<small>النتيجة النهائية السنوية</small>'
                            f'<br>{annual_val}</div>'
                        )
                    break
        except Exception:
            pass

        # ── الجدول الرئيسي ────────────────────────────────────
        parts.append(parse_table_from_ws(ws))

        parts.append('</div>')   # .page-wrapper
        first_page = False

    parts.append("</body></html>")
    return "".join(parts)

